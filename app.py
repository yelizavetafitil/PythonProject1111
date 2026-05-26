import sqlite3
import os
import collections
import re
import json
import hmac
import hashlib
import time
import threading
import tempfile
import sys
import subprocess
from contextlib import contextmanager
import calendar
import smtplib
from email.message import EmailMessage
from email.utils import format_datetime
from email.header import Header
from email import policy
from urllib.parse import urlparse
from urllib.parse import urlencode
from urllib import request as urllib_request
from urllib.error import URLError, HTTPError
from datetime import datetime
from datetime import timedelta
from datetime import date as date_type
from zoneinfo import ZoneInfo
from flask import (
    Flask,
    Response,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
import pandas as pd
import openpyxl
try:
    import xlrd
except Exception:
    xlrd = None

from tabel_fs import (
    configure_tabel_locale,
    join_tabel_path,
    linux_unc_misconfiguration,
    listdir_tabel,
    normalize_tabel_path,
    tabel_default_base_dir,
    tabel_default_leaders_file,
    tabel_path_exists,
    tabel_path_status,
    walk_tabel_excel,
)

configure_tabel_locale()

# --- ЗАПЛАТКА ДЛЯ LDAP3 ---
if not hasattr(collections, 'MutableMapping'):
    import collections.abc

    collections.MutableMapping = collections.abc.MutableMapping
if not hasattr(collections, 'Sequence'):
    import collections.abc

    collections.Sequence = collections.abc.Sequence

from ldap3 import Server, Connection, ALL, SUBTREE
from ldap3.core.exceptions import LDAPException

app = Flask(__name__)
app.secret_key = 'enterprise_hub_production_v37'
app.config.update(
    SESSION_COOKIE_NAME='enterprise_hub_session',
    SESSION_COOKIE_PATH='/',
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    SESSION_REFRESH_EACH_REQUEST=False
)

LDAP_CONFIG = {
    'uri': "ldap://192.168.0.4",
    'base': "DC=local,DC=energoprom,DC=by",
    'bind_dn': "CN=OC1,OU=ОЦ,DC=local,DC=energoprom,DC=by",
    'bind_password': "Pass_OC_5678",
    'user_attr': "sAMAccountName"
}

MASTER_ADMINS = ['rapeiko', 'oc1']
DEFAULT_USERS_GROUP_NAME = 'Пользователи'
DB_PATH = 'database.db'
GYM_ROOM_NAME = 'Спортзал'
LOGO_FILENAME = 'image2_hq.png'
PHONEBOOK_PATH = 'phonebook.xlsx'
PHONEBOOK_COLUMNS = ['dept', 'pos', 'surname', 'name', 'work', 'home', 'mobile']
MAIL_DOMAIN = 'energoprom.by'
MAIL_SENDER = f'robot-bnp@{MAIL_DOMAIN}'
SMTP_HOST = os.environ.get('SMTP_HOST', '192.168.0.28')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USERNAME = os.environ.get('SMTP_USERNAME', 'robot-bnp')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '^nO@u(Flu5+zH&v>')
SMTP_USE_TLS = os.environ.get('SMTP_USE_TLS', '1') == '1'
APP_TZ = ZoneInfo('Europe/Minsk')
_phonebook_cache = None
_phonebook_mtime = None
_ad_groups_cache = {}
TABEL_BASE_DIR = os.environ.get('TABEL_BASE_DIR', tabel_default_base_dir())
TABEL_LEADERS_FILE = os.environ.get('TABEL_LEADERS_FILE', tabel_default_leaders_file())
TABEL_SHEET_NAME = os.environ.get('TABEL_SHEET_NAME', 'Табель_3')
TABEL_CACHE_FILE = os.environ.get('TABEL_CACHE_FILE', 'tabel_portal_cache.json')
TABEL_FILE_PATTERN = re.compile(r'^(.+?)_(\d{2})_(\d{2})\.xlsx?$', re.IGNORECASE)
TABEL_FIO_RE = re.compile(r'^[А-ЯЁ][а-яё\-]+\s+[А-ЯЁ]\s*\.\s*[А-ЯЁ]\s*\.\s*$', re.IGNORECASE)
TABEL_STATUS_MAP = {
    'Б': 'Больничный', 'Н': 'Неявки', 'Р': 'Роды', 'ОЖ': 'Уход за ребенком',
    'О': 'Отпуск', 'ЛО': 'Социальный отпуск', 'А': 'Отпуск без сохранения', 'Т': 'Нетрудоспособность',
    'Х': 'Уход за больным', 'ПР': 'Прогул', 'Г': 'Гособязанность', 'ОА': 'Инициатива нанимателя',
    'Д': 'Донор', 'У': 'Учеба', 'УД': 'Учебные дни', 'ДМ': 'День матери',
    'ДСП': 'Диспансеризация', 'П': 'Праздник', 'К': 'Командировка', 'В': 'Выходной'
}
TABEL_INDEX_LOCK = threading.Lock()
TABEL_INDEX = {}
TABEL_FILE_CACHE = {}
TABEL_LEADERS_CACHE = {}
TABEL_LEADERS_MTIME = 0.0
TABEL_LAST_SCAN_TS = 0.0
TABEL_SCAN_INTERVAL_SEC = 180
TABEL_LAST_SCAN_ERRORS = []
TABEL_LAST_SCAN_STATS = {}
TABEL_SMB_USER = os.environ.get('TABEL_SMB_USER', 'oc1@local.energoprom.by')
TABEL_SMB_PASSWORD = os.environ.get('TABEL_SMB_PASSWORD', LDAP_CONFIG['bind_password'])
TABEL_SMB_ENABLED = os.environ.get(
    'TABEL_SMB_ENABLED',
    '1' if sys.platform == 'win32' else '0',
).strip().lower() in ('1', 'true', 'yes')
TABEL_LAST_SMB_CONNECT = {}
TABEL_SMB_NET_USE_TIMEOUT = int(os.environ.get('TABEL_SMB_NET_USE_TIMEOUT', '20'))
TABEL_SCAN_LOCK = threading.Lock()
TABEL_SCAN_THREAD = None
_smb_use_lock = threading.Lock()
KNOWLEDGE_BASE_ROOT = os.environ.get('KNOWLEDGE_BASE_ROOT', os.path.join(app.root_path, 'БАЗА ЗНАНИЙ'))
KNOWLEDGE_BASE_INSTRUCTIONS_DIR = os.environ.get(
    'KNOWLEDGE_BASE_INSTRUCTIONS_DIR',
    KNOWLEDGE_BASE_ROOT
)
AI_ASSISTANT_URL = os.environ.get('AI_ASSISTANT_URL', 'http://localhost:5000/sso-login')
AI_SSO_SHARED_SECRET = os.environ.get('AI_SSO_SHARED_SECRET', 'change-this-ai-sso-secret')


def normalize_resource_url(raw_url):
    value = (raw_url or '').strip()
    if not value:
        return value
    try:
        parsed = urlparse(value)
    except Exception:
        return value
    host = (parsed.hostname or '').lower()
    if host in ('127.0.0.1', 'localhost', '::1'):
        path = parsed.path or '/'
        query = f'?{parsed.query}' if parsed.query else ''
        fragment = f'#{parsed.fragment}' if parsed.fragment else ''
        return f'{path}{query}{fragment}'
    return value


def normalize_ad_username(raw_username):
    username = (raw_username or '').strip().lower()
    if not username:
        return ''
    if '\\' in username:
        username = username.split('\\')[-1].strip()
    if '@' in username:
        username = username.split('@')[0].strip()
    return username


def _tabel_clean_fio(raw_val):
    if pd.isna(raw_val):
        return ''
    text = str(raw_val).strip()
    return text.replace('\xa0', ' ').replace('\u200b', '')


def _tabel_is_work_value(value_str):
    if value_str in ('', '0', '0.0'):
        return False
    return value_str.replace('.', '', 1).replace(',', '', 1).isdigit()


def _tabel_parse_filename(filename):
    match = TABEL_FILE_PATTERN.match(filename or '')
    if not match:
        return None, None
    return match.group(2), match.group(3)


def _tabel_read_any_excel(path):
    path = normalize_tabel_path(path)
    if not path or not tabel_path_exists(path):
        return None
    try:
        if path.lower().endswith('.xlsx'):
            return pd.read_excel(path, sheet_name=TABEL_SHEET_NAME, engine='openpyxl', header=None)
        if xlrd is None:
            return None
        workbook = xlrd.open_workbook(path, formatting_info=False)
        sheet = workbook.sheet_by_name(TABEL_SHEET_NAME)
        return pd.DataFrame([sheet.row_values(i) for i in range(sheet.nrows)])
    except Exception:
        return None


def _tabel_load_cache():
    global TABEL_FILE_CACHE
    if not os.path.exists(TABEL_CACHE_FILE):
        TABEL_FILE_CACHE = {}
        return
    try:
        with open(TABEL_CACHE_FILE, 'r', encoding='utf-8') as cache_file:
            parsed = json.load(cache_file)
            TABEL_FILE_CACHE = parsed if isinstance(parsed, dict) else {}
    except Exception:
        TABEL_FILE_CACHE = {}


def _tabel_save_cache():
    try:
        with open(TABEL_CACHE_FILE, 'w', encoding='utf-8') as cache_file:
            json.dump(TABEL_FILE_CACHE, cache_file, ensure_ascii=False)
    except Exception:
        return


def _tabel_smb_configured():
    return bool(TABEL_SMB_ENABLED and TABEL_SMB_USER and TABEL_SMB_PASSWORD)


def _windows_net_use(unc_path, username, password, *, disconnect=False):
    """Подключение/отключение UNC на Windows (net use), как в старом сервисе табеля."""
    if sys.platform != 'win32':
        return False, 'Только Windows'
    if disconnect:
        subprocess.run(
            ['net', 'use', unc_path, '/delete', '/y'],
            capture_output=True,
            text=True,
            timeout=30,
        )
        return True, ''
    cmd = ['net', 'use', unc_path, password, f'/user:{username}', '/persistent:no']
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=TABEL_SMB_NET_USE_TIMEOUT)
    output = ((result.stdout or '') + (result.stderr or '')).strip()
    if result.returncode == 0:
        return True, output
    lowered = output.lower()
    if '1219' in output or 'уже подключ' in lowered or 'already' in lowered:
        return True, output
    if '1326' in output or '86' in output or 'denied' in lowered or 'отказано' in lowered:
        return False, output or 'Неверный логин или пароль / нет прав'
    subprocess.run(['net', 'use', unc_path, '/delete', '/y'], capture_output=True, text=True, timeout=30)
    retry = subprocess.run(cmd, capture_output=True, text=True, timeout=TABEL_SMB_NET_USE_TIMEOUT)
    output2 = ((retry.stdout or '') + (retry.stderr or '')).strip()
    if retry.returncode == 0:
        return True, output2
    return False, output2 or output or f'net use код {retry.returncode}'


@contextmanager
def _tabel_unc_connection():
    """Windows: net use. Linux: чтение с mount (пути /mnt/tabel, см. tabel_fs)."""
    global TABEL_LAST_SMB_CONNECT
    linux_misconfig = linux_unc_misconfiguration(TABEL_BASE_DIR, TABEL_LEADERS_FILE)
    if linux_misconfig:
        TABEL_LAST_SMB_CONNECT = {'connected': False, 'user': None, 'message': linux_misconfig}
        yield TABEL_LAST_SMB_CONNECT
        return
    if sys.platform != 'win32':
        base = normalize_tabel_path(TABEL_BASE_DIR)
        ok = tabel_path_exists(base)
        TABEL_LAST_SMB_CONNECT = {
            'connected': ok,
            'user': None,
            'message': f'Linux mount: {base}' if ok else f'Нет каталога {base} (смонтируйте //srv-doc/ТАБЕЛЬ)',
        }
        yield TABEL_LAST_SMB_CONNECT
        return
    if not _tabel_smb_configured():
        TABEL_LAST_SMB_CONNECT = {'connected': False, 'user': None, 'message': 'Отключено (TABEL_SMB_ENABLED=0)'}
        yield TABEL_LAST_SMB_CONNECT
        return
    unc = TABEL_BASE_DIR
    with _smb_use_lock:
        ok, message = _windows_net_use(unc, TABEL_SMB_USER, TABEL_SMB_PASSWORD)
        TABEL_LAST_SMB_CONNECT = {
            'connected': ok,
            'user': TABEL_SMB_USER,
            'message': message,
        }
        try:
            yield TABEL_LAST_SMB_CONNECT
        finally:
            if ok:
                _windows_net_use(unc, TABEL_SMB_USER, TABEL_SMB_PASSWORD, disconnect=True)


def _tabel_rebuild_index_from_cache():
    local_index = {}
    for file_path, data in TABEL_FILE_CACHE.items():
        if not isinstance(data, dict):
            continue
        dept = data.get('dept')
        yy_mm = data.get('yy_mm')
        emps = data.get('emps')
        if not dept or not yy_mm or not isinstance(emps, list):
            continue
        local_index.setdefault(dept, {}).setdefault(yy_mm, []).append({'file': file_path, 'employees': emps})
    with TABEL_INDEX_LOCK:
        TABEL_INDEX.clear()
        TABEL_INDEX.update(local_index)


def _scan_tabel_base_dir():
    global TABEL_FILE_CACHE, TABEL_LAST_SCAN_ERRORS, TABEL_LAST_SCAN_STATS
    current_year = datetime.now().year
    found_files = set()
    cache_updated = False
    errors = []
    stats = {
        'files_matched': 0,
        'files_parsed': 0,
        'files_skipped_cache': 0,
        'files_failed': 0,
        'employees_indexed': 0,
    }

    with _tabel_unc_connection() as smb:
        if linux_unc_misconfiguration(TABEL_BASE_DIR, TABEL_LEADERS_FILE):
            errors.append({
                'stage': 'linux_paths',
                'path': TABEL_BASE_DIR,
                'message': linux_unc_misconfiguration(TABEL_BASE_DIR, TABEL_LEADERS_FILE),
            })
        smb_ready = not _tabel_smb_configured() or bool(smb.get('connected'))
        if sys.platform != 'win32':
            smb_ready = bool(smb.get('connected'))
        if _tabel_smb_configured() and not smb_ready:
            errors.append({
                'stage': 'smb_connect',
                'path': TABEL_BASE_DIR,
                'message': smb.get('message') or f'Не удалось подключить шару под {TABEL_SMB_USER}',
            })

        if not smb_ready:
            _tabel_rebuild_index_from_cache()
            TABEL_LAST_SCAN_ERRORS = errors[-100:]
            TABEL_LAST_SCAN_STATS = stats
            return

        base_dir = normalize_tabel_path(TABEL_BASE_DIR)
        if not tabel_path_exists(base_dir):
            errors.append({
                'stage': 'base_dir',
                'path': base_dir,
                'message': 'Каталог не найден или недоступен по сети',
            })
        else:
            try:
                listdir_tabel(base_dir)
            except OSError as exc:
                errors.append({
                    'stage': 'base_dir',
                    'path': base_dir,
                    'message': f'Нет прав на чтение каталога: {exc}',
                })

        try:
            for root, dept, filename in walk_tabel_excel(base_dir):
                mm, yy = _tabel_parse_filename(filename)
                if not mm:
                    continue
                try:
                    if 2000 + int(yy) < current_year - 1:
                        continue
                except Exception:
                    continue
                full_path = join_tabel_path(root, filename)
                found_files.add(full_path)
                stats['files_matched'] += 1
                try:
                    current_mtime = os.path.getmtime(full_path)
                except OSError as exc:
                    stats['files_failed'] += 1
                    errors.append({
                        'stage': 'mtime',
                        'path': full_path,
                        'message': str(exc),
                    })
                    continue
                cached = TABEL_FILE_CACHE.get(full_path)
                if isinstance(cached, dict) and cached.get('mtime') == current_mtime:
                    stats['files_skipped_cache'] += 1
                    stats['employees_indexed'] += len(cached.get('emps') or [])
                    continue
                df = _tabel_read_any_excel(full_path)
                if df is None:
                    stats['files_failed'] += 1
                    errors.append({
                        'stage': 'read_excel',
                        'path': full_path,
                        'message': f'Не удалось прочитать лист «{TABEL_SHEET_NAME}»',
                    })
                    continue
                employees = []
                try:
                    for row_index, raw_name in enumerate(df.iloc[:, 1]):
                        name = _tabel_clean_fio(raw_name)
                        if not name or not TABEL_FIO_RE.match(name):
                            continue
                        row_data = df.iloc[row_index].values
                        days_data = []
                        for day_idx in range(31):
                            col_idx = 2 + day_idx
                            if col_idx < len(row_data):
                                day_val = row_data[col_idx]
                                days_data.append(str(day_val).strip() if not pd.isna(day_val) else '')
                            else:
                                days_data.append('')
                        employees.append({'fio': name, 'days': days_data})
                except Exception as exc:
                    stats['files_failed'] += 1
                    errors.append({
                        'stage': 'parse_rows',
                        'path': full_path,
                        'message': str(exc),
                    })
                    continue
                if employees:
                    TABEL_FILE_CACHE[full_path] = {
                        'mtime': current_mtime,
                        'dept': dept,
                        'yy_mm': f'{yy}_{mm}',
                        'emps': employees
                    }
                    cache_updated = True
                    stats['files_parsed'] += 1
                    stats['employees_indexed'] += len(employees)
                else:
                    stats['files_failed'] += 1
                    errors.append({
                        'stage': 'empty_sheet',
                        'path': full_path,
                        'message': 'В файле нет строк с ФИО в ожидаемом формате',
                    })
        except OSError as exc:
            errors.append({
                'stage': 'walk',
                'path': TABEL_BASE_DIR,
                'message': str(exc),
            })

        leaders_path = normalize_tabel_path(TABEL_LEADERS_FILE)
        if not tabel_path_exists(leaders_path):
            errors.append({
                'stage': 'leaders_file',
                'path': leaders_path,
                'message': 'Файл списка руководителей не найден',
            })
        else:
            try:
                openpyxl.load_workbook(leaders_path, read_only=True, data_only=True).close()
            except Exception as exc:
                errors.append({
                    'stage': 'leaders_file',
                    'path': TABEL_LEADERS_FILE,
                    'message': f'Не удалось открыть: {exc}',
                })

        deleted_files = set(TABEL_FILE_CACHE.keys()) - found_files
        if deleted_files:
            for deleted_path in deleted_files:
                TABEL_FILE_CACHE.pop(deleted_path, None)
            cache_updated = True
        if cache_updated:
            _tabel_save_cache()
        _tabel_rebuild_index_from_cache()
        TABEL_LAST_SCAN_ERRORS = errors[-100:]
        TABEL_LAST_SCAN_STATS = stats


def _run_tabel_scan():
    global TABEL_LAST_SCAN_TS
    _scan_tabel_base_dir()
    TABEL_LAST_SCAN_TS = time.time()


def _schedule_tabel_scan(force=False):
    global TABEL_SCAN_THREAD

    def _worker():
        if not TABEL_SCAN_LOCK.acquire(blocking=False):
            return
        try:
            now_ts = time.time()
            with TABEL_INDEX_LOCK:
                has_index = bool(TABEL_INDEX)
            if not force and (now_ts - TABEL_LAST_SCAN_TS) < TABEL_SCAN_INTERVAL_SEC and has_index:
                return
            _run_tabel_scan()
        finally:
            TABEL_SCAN_LOCK.release()

    if TABEL_SCAN_THREAD and TABEL_SCAN_THREAD.is_alive():
        return
    TABEL_SCAN_THREAD = threading.Thread(target=_worker, name='tabel-scan', daemon=True)
    TABEL_SCAN_THREAD.start()


def ensure_tabel_index(force=False, blocking=False):
    now_ts = time.time()
    with TABEL_INDEX_LOCK:
        has_index = bool(TABEL_INDEX)
    if not force and (now_ts - TABEL_LAST_SCAN_TS) < TABEL_SCAN_INTERVAL_SEC and has_index:
        return
    if blocking or force:
        with TABEL_SCAN_LOCK:
            _run_tabel_scan()
        return
    _schedule_tabel_scan(force=force)


def _tabel_index_summary():
    departments = []
    periods = set()
    employees = set()
    with TABEL_INDEX_LOCK:
        for dept_name, dep_periods in TABEL_INDEX.items():
            dept_periods_list = sorted(dep_periods.keys())
            periods.update(dep_periods_list)
            dept_employees = set()
            for period_key, records in dep_periods.items():
                for record in records:
                    for emp in record.get('employees', []):
                        fio = emp.get('fio')
                        if fio:
                            employees.add(fio)
                            dept_employees.add(fio)
            departments.append({
                'name': dept_name,
                'periods': dept_periods_list,
                'employees': len(dept_employees),
            })
    return {
        'departments': len(departments),
        'department_list': departments,
        'periods': sorted(periods, reverse=True),
        'unique_employees': len(employees),
    }


def _tabel_get_current_status(fio):
    now = datetime.now()
    curr_period = f'{str(now.year)[2:]}_{now.month:02d}'
    day_idx = now.day - 1
    with TABEL_INDEX_LOCK:
        for dep_data in TABEL_INDEX.values():
            records = dep_data.get(curr_period) or []
            for rec in records:
                match = next((item for item in rec.get('employees', []) if item.get('fio') == fio), None)
                if not match:
                    continue
                days = match.get('days') or []
                if day_idx >= len(days):
                    return 'unknown'
                val = str(days[day_idx]).strip().upper()
                if val == '':
                    return 'absent'
                if _tabel_is_work_value(val):
                    return 'work'
                if val == 'В':
                    return 'rest'
                return 'absent'
    return 'unknown'


def get_tabel_leaders_data(use_network=True):
    global TABEL_LEADERS_CACHE, TABEL_LEADERS_MTIME
    categories = [
        'Руководство',
        'Персонал при руководстве',
        'Производственные отделы',
        'Главные инженеры проекта',
        'Непроизводственные отделы'
    ]
    try:
        leaders_path = normalize_tabel_path(TABEL_LEADERS_FILE)
        current_mtime = os.path.getmtime(leaders_path) if tabel_path_exists(leaders_path) else 0
    except Exception:
        current_mtime = 0
    if TABEL_LEADERS_CACHE and current_mtime == TABEL_LEADERS_MTIME:
        for category in categories:
            for leader in TABEL_LEADERS_CACHE.get(category, []):
                leader['status_cls'] = f"st-{_tabel_get_current_status(leader.get('fio', ''))}"
        return TABEL_LEADERS_CACHE
    data = {category: [] for category in categories}
    if not use_network:
        return data if not TABEL_LEADERS_CACHE else TABEL_LEADERS_CACHE
    with _tabel_unc_connection():
        leaders_path = normalize_tabel_path(TABEL_LEADERS_FILE)
        if not tabel_path_exists(leaders_path):
            return data
        try:
            workbook = openpyxl.load_workbook(leaders_path, data_only=True)
            sheet = workbook.active
            current_category = None
            for row in sheet.iter_rows(values_only=True):
                first_val = _tabel_clean_fio(row[0] if row else '')
                if not first_val:
                    continue
                matched_category = next((cat for cat in categories if cat.lower() == first_val.lower()), None)
                if matched_category:
                    current_category = matched_category
                    continue
                if current_category and TABEL_FIO_RE.match(first_val):
                    status = _tabel_get_current_status(first_val)
                    data[current_category].append({
                        'fio': first_val,
                        'info': f"{str((row[1] if len(row) > 1 else '') or '')} {str((row[2] if len(row) > 2 else '') or '')}".strip(),
                        'status_cls': f'st-{status}'
                    })
            TABEL_LEADERS_CACHE = data
            TABEL_LEADERS_MTIME = current_mtime
        except Exception:
            return data
    return data


def get_db_connection():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.execute('PRAGMA journal_mode=WAL;')
    conn.row_factory = sqlite3.Row
    return conn


def get_users_group_id(conn):
    row = conn.execute(
        'SELECT id FROM groups WHERE name = ? COLLATE NOCASE',
        (DEFAULT_USERS_GROUP_NAME,),
    ).fetchone()
    return row['id'] if row else None


def ensure_default_users_group(conn):
    conn.execute('INSERT OR IGNORE INTO groups (name) VALUES (?)', (DEFAULT_USERS_GROUP_NAME,))


def add_user_to_default_group(conn, username):
    login = normalize_ad_username(username)
    if not login:
        return
    ensure_default_users_group(conn)
    group_id = get_users_group_id(conn)
    if not group_id:
        return
    conn.execute(
        'INSERT OR IGNORE INTO group_members (group_id, username) VALUES (?, ?)',
        (group_id, login),
    )


def fetch_all_ad_usernames():
    usernames = set()
    server = Server(LDAP_CONFIG['uri'], get_info=ALL, connect_timeout=10)
    conn = Connection(
        server,
        user=LDAP_CONFIG['bind_dn'],
        password=LDAP_CONFIG['bind_password'],
        auto_bind=True,
    )
    search_filter = '(&(objectCategory=person)(objectClass=user)(sAMAccountName=*))'
    conn.search(
        LDAP_CONFIG['base'],
        search_filter,
        SUBTREE,
        attributes=['sAMAccountName'],
    )
    for entry in conn.entries:
        if entry.sAMAccountName:
            login = normalize_ad_username(str(entry.sAMAccountName))
            if login:
                usernames.add(login)
    return sorted(usernames)


def sync_all_ad_users_to_default_group(conn):
    ensure_default_users_group(conn)
    group_id = get_users_group_id(conn)
    if not group_id:
        return 0
    try:
        usernames = fetch_all_ad_usernames()
    except Exception as exc:
        app.logger.warning('AD sync for default users group failed: %s', exc)
        return 0
    added = 0
    for login in usernames:
        cur = conn.execute(
            'INSERT OR IGNORE INTO group_members (group_id, username) VALUES (?, ?)',
            (group_id, login),
        )
        if cur.rowcount:
            added += 1
    return added


def ensure_gym_room_exists():
    with get_db_connection() as conn:
        conn.execute('INSERT OR IGNORE INTO meeting_rooms (name) VALUES (?)', (GYM_ROOM_NAME,))
        conn.commit()


def format_phone(phone):
    source = str(phone).replace('.0', '').strip()
    digits = re.sub(r'\D', '', source)
    if len(digits) == 12:
        return f"+{digits[:3]}-{digits[3:5]}-{digits[5:8]}-{digits[8:10]}-{digits[10:]}"
    return source


def get_phonebook_contacts():
    global _phonebook_cache, _phonebook_mtime
    if not os.path.exists(PHONEBOOK_PATH):
        return []

    current_mtime = os.path.getmtime(PHONEBOOK_PATH)
    if _phonebook_cache is not None and _phonebook_mtime == current_mtime:
        return _phonebook_cache

    df = pd.read_excel(PHONEBOOK_PATH, names=PHONEBOOK_COLUMNS)
    df['work'] = df['work'].apply(format_phone)
    df['mobile'] = df['mobile'].apply(format_phone)
    df['home'] = df['home'].apply(format_phone)

    _phonebook_cache = df.fillna('').to_dict(orient='records')
    _phonebook_mtime = current_mtime
    return _phonebook_cache


# Внутренние URL портала → таблица «расширенного доступа» в /manage (скрин 2).
RESOURCE_PATH_PRIVILEGE_TABLE = {
    '/tabel': 'tabel_privileged_entities',
    '/phonebook': 'phonebook_privileged_entities',
    '/meeting-rooms': 'booking_privileged_entities',
    '/gym-booking': 'booking_privileged_entities',
    '/driver-trips': 'booking_privileged_entities',
    '/ai-assistant': 'ai_privileged_entities',
}


def _resource_internal_path(url):
    normalized = normalize_resource_url(url or '')
    if not normalized:
        return None
    raw = normalized.split('?')[0].split('#')[0]
    if raw.startswith('/'):
        return raw.rstrip('/') or '/'
    try:
        parsed = urlparse(normalized)
        if not parsed.netloc:
            return None
        host = (parsed.hostname or '').lower()
        if host in ('127.0.0.1', 'localhost', '::1'):
            p = parsed.path or '/'
            return p.rstrip('/') or '/'
        return None
    except Exception:
        return None


def _load_resource_group_access_maps(conn):
    members_rows = conn.execute('SELECT group_id, username FROM group_members').fetchall()
    group_map = {}
    for mr in members_rows:
        gid = str(mr['group_id'])
        group_map.setdefault(gid, []).append(normalize_ad_username(mr['username']))
    users_group_row = conn.execute(
        'SELECT id FROM groups WHERE name = ? COLLATE NOCASE',
        (DEFAULT_USERS_GROUP_NAME,),
    ).fetchone()
    users_group_id = str(users_group_row['id']) if users_group_row else None
    return group_map, users_group_id


def _user_matches_resource_groups(username, group_ids, group_map, users_group_id, user_ad_groups):
    login = normalize_ad_username(username)
    if not group_ids:
        return False
    ad_lower = {g.strip().lower() for g in user_ad_groups if g}
    for gid in group_ids:
        if users_group_id and gid == users_group_id:
            return True
        allowed_entities = group_map.get(gid, [])
        if login in allowed_entities or any(ag in allowed_entities for ag in ad_lower):
            return True
    return False


def _user_has_group_access_to_resource_url(username, url_path, conn=None):
    url_path = (url_path or '').rstrip('/') or '/'
    login = normalize_ad_username(username)
    if not login:
        return False
    own_conn = conn is None
    if own_conn:
        conn = get_db_connection()
    try:
        user_ad_groups = get_user_ad_groups_by_username(login)
        group_map, users_group_id = _load_resource_group_access_maps(conn)
        rows = conn.execute('''
            SELECT r.url, GROUP_CONCAT(ga.group_id) as group_ids
            FROM resources r
            LEFT JOIN resource_group_access ga ON r.id = ga.resource_id
            GROUP BY r.id
        ''').fetchall()
        for row in rows:
            path = _resource_internal_path(row['url'])
            if path != url_path:
                continue
            g_ids = [g for g in (row['group_ids'] or '').split(',') if g]
            if not g_ids:
                return True
            if _user_matches_resource_groups(login, g_ids, group_map, users_group_id, user_ad_groups):
                return True
        return False
    finally:
        if own_conn:
            conn.close()


def can_access_portal_path(username, path):
    """Доступ к странице: группа на карточке ресурса (скрин 1) ИЛИ запись в /manage (скрин 2)."""
    login = normalize_ad_username(username)
    if not login:
        return False
    if login in MASTER_ADMINS:
        return True
    path = (path or '').rstrip('/') or '/'
    table = RESOURCE_PATH_PRIVILEGE_TABLE.get(path)
    if table and _has_privileged_entity_access(login, table):
        return True
    return _user_has_group_access_to_resource_url(login, path)


def can_access_portal_resource(username, url, group_ids_str, group_map, users_group_id, user_ad_groups):
    """Видимость плитки на главной — та же логика, что и can_access_portal_path."""
    login = normalize_ad_username(username)
    if not login:
        return False
    if login in MASTER_ADMINS:
        return True
    g_ids = [g for g in (group_ids_str or '').split(',') if g]
    has_group = (
        _user_matches_resource_groups(login, g_ids, group_map, users_group_id, user_ad_groups)
        if g_ids else False
    )
    path = _resource_internal_path(normalize_resource_url(url))
    priv_table = RESOURCE_PATH_PRIVILEGE_TABLE.get(path) if path else None
    has_priv = _has_privileged_entity_access(login, priv_table) if priv_table else False
    if not g_ids:
        if priv_table:
            return has_priv
        return True
    return has_group or has_priv


def can_view_extended_phonebook(username):
    return can_access_portal_path(username, '/phonebook')


def _has_privileged_entity_access(login, table_name):
    user_groups = {group.strip().lower() for group in get_user_ad_groups_by_username(login)}
    conn = get_db_connection()
    try:
        entity_rows = conn.execute(
            f'SELECT entity_type, entity_login FROM {table_name}'
        ).fetchall()
    finally:
        conn.close()
    if not entity_rows:
        return False
    allowed_users = set()
    allowed_groups = set()
    for row in entity_rows:
        entity_type = (row['entity_type'] or '').strip().lower()
        entity_login = (row['entity_login'] or '').strip().lower()
        if not entity_login:
            continue
        if entity_type == 'group':
            allowed_groups.add(entity_login)
        else:
            allowed_users.add(entity_login)
    if login in allowed_users:
        return True
    if user_groups.intersection(allowed_groups):
        return True
    return False


def can_manage_all_bookings(username):
    login = normalize_ad_username(username)
    if not login:
        return False
    if login in MASTER_ADMINS:
        return True
    return _has_privileged_entity_access(login, 'booking_privileged_entities')


def can_manage_resources(username):
    login = normalize_ad_username(username)
    if not login:
        return False
    if login in MASTER_ADMINS:
        return True
    return _has_privileged_entity_access(login, 'resource_privileged_entities')


def can_use_ai_assistant(username):
    return can_access_portal_path(username, '/ai-assistant')


def can_view_tabel(username):
    return can_access_portal_path(username, '/tabel')


def _build_ai_sso_url(username, display_name):
    login = normalize_ad_username(username)
    display = (display_name or login or '').strip()
    ts = str(int(time.time()))
    payload = f'{login}|{display}|{ts}'
    signature = hmac.new(
        AI_SSO_SHARED_SECRET.encode('utf-8'),
        payload.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    query = urlencode({'u': login, 'd': display, 'ts': ts, 'sig': signature})
    separator = '&' if '?' in AI_ASSISTANT_URL else '?'
    return f'{AI_ASSISTANT_URL}{separator}{query}'


def _knowledge_base_collect_categories():
    categories = []
    base_path = os.path.realpath(KNOWLEDGE_BASE_INSTRUCTIONS_DIR)
    if not os.path.isdir(base_path):
        return categories
    for name in os.listdir(base_path):
        full_path = os.path.join(base_path, name)
        if not os.path.isdir(full_path):
            continue
        categories.append(name)
    return sorted(categories, key=lambda item: item.lower())


def _knowledge_base_resolve_category_path(category_name):
    base_path = os.path.realpath(KNOWLEDGE_BASE_INSTRUCTIONS_DIR)
    category_path = os.path.realpath(os.path.join(base_path, category_name))
    if not category_path.startswith(base_path + os.sep):
        return None
    if not os.path.isdir(category_path):
        return None
    return category_path


def _knowledge_base_collect_files(category_name):
    category_path = _knowledge_base_resolve_category_path(category_name)
    if not category_path:
        return []
    file_items = []
    for root, _, files in os.walk(category_path):
        for filename in files:
            if not filename.lower().endswith('.pdf'):
                continue
            full_path = os.path.join(root, filename)
            rel_path = os.path.relpath(full_path, category_path).replace('\\', '/')
            file_items.append({
                'name': filename,
                'path': rel_path
            })
    return sorted(file_items, key=lambda item: item['path'].lower())


def _knowledge_base_collect_all_files():
    items = []
    for category in _knowledge_base_collect_categories():
        for file_item in _knowledge_base_collect_files(category):
            items.append({
                'category': category,
                'name': file_item.get('name', ''),
                'path': file_item.get('path', '')
            })
    return items


def _knowledge_base_resolve_file_path(category_name, relative_file_path):
    category_path = _knowledge_base_resolve_category_path(category_name)
    if not category_path:
        return None
    normalized_rel = (relative_file_path or '').replace('\\', '/').strip('/')
    if not normalized_rel or not normalized_rel.lower().endswith('.pdf'):
        return None
    target_path = os.path.realpath(os.path.join(category_path, normalized_rel))
    if not target_path.startswith(category_path + os.sep):
        return None
    if not os.path.isfile(target_path):
        return None
    return target_path


def init_db():
    with get_db_connection() as conn:
        # 1. Проверяем/Обновляем таблицу ресурсов (удаляем старый столбец access_group_id если он мешает)
        conn.execute('''
            CREATE TABLE IF NOT EXISTS resources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL, url TEXT NOT NULL,
                category TEXT NOT NULL, desc TEXT, 
                position INTEGER DEFAULT 0
            )''')

        # 2. Создаем новую таблицу связей (МНОГИЕ-КО-МНОГИМ)
        conn.execute('''
            CREATE TABLE IF NOT EXISTS resource_group_access (
                resource_id INTEGER,
                group_id INTEGER,
                PRIMARY KEY (resource_id, group_id)
            )''')

        # 3. Группы
        conn.execute('CREATE TABLE IF NOT EXISTS groups (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE)')

        # 3.1 Разделы (категории ресурсов)
        conn.execute('CREATE TABLE IF NOT EXISTS categories (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE)')

        # 4. Участники
        conn.execute('''
            CREATE TABLE IF NOT EXISTS group_members (
                group_id INTEGER, username TEXT, 
                PRIMARY KEY (group_id, username)
            )''')
        # 5. Переговорки
        conn.execute('''
            CREATE TABLE IF NOT EXISTS meeting_rooms (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )''')

        # 6. Брони переговорок
        conn.execute('''
            CREATE TABLE IF NOT EXISTS meeting_bookings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                room_id INTEGER NOT NULL,
                booked_by TEXT NOT NULL,
                purpose TEXT NOT NULL,
                participants_json TEXT DEFAULT '[]',
                meeting_date TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                booking_status TEXT DEFAULT 'active',
                canceled_by TEXT,
                canceled_at TEXT,
                owner_username TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(room_id) REFERENCES meeting_rooms(id)
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS meeting_booking_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                booking_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                changed_by TEXT NOT NULL,
                changed_at TEXT NOT NULL,
                details_json TEXT NOT NULL,
                FOREIGN KEY(booking_id) REFERENCES meeting_bookings(id)
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS driver_trips (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vehicle_model TEXT NOT NULL,
                vehicle_color TEXT NOT NULL DEFAULT '#1f77b4',
                trip_date TEXT NOT NULL,
                departure_time TEXT NOT NULL,
                origin TEXT NOT NULL DEFAULT 'РУП «Белнипиэнергопром»',
                route_stops TEXT DEFAULT '',
                destination TEXT NOT NULL,
                description TEXT,
                trip_status TEXT NOT NULL DEFAULT 'active',
                canceled_by TEXT,
                canceled_at TEXT,
                owner_username TEXT,
                created_by TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS driver_trip_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trip_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                changed_by TEXT NOT NULL,
                changed_at TEXT NOT NULL,
                details_json TEXT NOT NULL,
                FOREIGN KEY(trip_id) REFERENCES driver_trips(id)
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS phonebook_privileged_users (
                username TEXT PRIMARY KEY
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS phonebook_privileged_entities (
                entity_type TEXT NOT NULL,
                entity_login TEXT NOT NULL,
                PRIMARY KEY (entity_type, entity_login)
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS booking_privileged_entities (
                entity_type TEXT NOT NULL,
                entity_login TEXT NOT NULL,
                PRIMARY KEY (entity_type, entity_login)
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS resource_privileged_entities (
                entity_type TEXT NOT NULL,
                entity_login TEXT NOT NULL,
                PRIMARY KEY (entity_type, entity_login)
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS ai_privileged_entities (
                entity_type TEXT NOT NULL,
                entity_login TEXT NOT NULL,
                PRIMARY KEY (entity_type, entity_login)
            )''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS tabel_privileged_entities (
                entity_type TEXT NOT NULL,
                entity_login TEXT NOT NULL,
                PRIMARY KEY (entity_type, entity_login)
            )''')
        # Миграция старого формата (только пользователи) в новый универсальный справочник.
        legacy_rows = conn.execute('SELECT username FROM phonebook_privileged_users').fetchall()
        for row in legacy_rows:
            legacy_login = normalize_ad_username(row['username'])
            if not legacy_login:
                continue
            conn.execute(
                'INSERT OR IGNORE INTO phonebook_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                ('user', legacy_login)
            )
        booking_columns = conn.execute("PRAGMA table_info(meeting_bookings)").fetchall()
        booking_column_names = {row['name'] for row in booking_columns}
        if 'participants_json' not in booking_column_names:
            conn.execute("ALTER TABLE meeting_bookings ADD COLUMN participants_json TEXT DEFAULT '[]'")
        if 'booking_status' not in booking_column_names:
            conn.execute("ALTER TABLE meeting_bookings ADD COLUMN booking_status TEXT DEFAULT 'active'")
            conn.execute("UPDATE meeting_bookings SET booking_status = 'active' WHERE booking_status IS NULL")
        if 'canceled_by' not in booking_column_names:
            conn.execute("ALTER TABLE meeting_bookings ADD COLUMN canceled_by TEXT")
        if 'canceled_at' not in booking_column_names:
            conn.execute("ALTER TABLE meeting_bookings ADD COLUMN canceled_at TEXT")
        driver_trip_columns = conn.execute("PRAGMA table_info(driver_trips)").fetchall()
        driver_trip_column_names = {row['name'] for row in driver_trip_columns}
        if 'vehicle_color' not in driver_trip_column_names:
            conn.execute("ALTER TABLE driver_trips ADD COLUMN vehicle_color TEXT NOT NULL DEFAULT '#1f77b4'")
        if 'trip_status' not in driver_trip_column_names:
            conn.execute("ALTER TABLE driver_trips ADD COLUMN trip_status TEXT NOT NULL DEFAULT 'active'")
            conn.execute("UPDATE driver_trips SET trip_status = 'active' WHERE trip_status IS NULL")
        if 'canceled_by' not in driver_trip_column_names:
            conn.execute("ALTER TABLE driver_trips ADD COLUMN canceled_by TEXT")
        if 'canceled_at' not in driver_trip_column_names:
            conn.execute("ALTER TABLE driver_trips ADD COLUMN canceled_at TEXT")
        if 'owner_username' not in driver_trip_column_names:
            conn.execute("ALTER TABLE driver_trips ADD COLUMN owner_username TEXT")
            conn.execute("UPDATE driver_trips SET owner_username = created_by WHERE owner_username IS NULL OR owner_username = ''")
        if 'updated_at' not in driver_trip_column_names:
            conn.execute("ALTER TABLE driver_trips ADD COLUMN updated_at TEXT")
            conn.execute("UPDATE driver_trips SET updated_at = CURRENT_TIMESTAMP WHERE updated_at IS NULL")
        if 'route_stops' not in driver_trip_column_names:
            conn.execute("ALTER TABLE driver_trips ADD COLUMN route_stops TEXT DEFAULT ''")
            conn.execute("UPDATE driver_trips SET route_stops = '' WHERE route_stops IS NULL")
        # Синхронизируем справочник разделов с уже существующими ресурсами
        existing_categories = conn.execute(
            "SELECT DISTINCT TRIM(category) AS category FROM resources WHERE TRIM(category) <> ''"
        ).fetchall()
        for row in existing_categories:
            conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (row['category'],))
        # Нормализуем внутренние URL ресурсов, чтобы проект корректно открывался на любом ПК.
        resource_rows = conn.execute('SELECT id, url FROM resources').fetchall()
        for row in resource_rows:
            normalized_url = normalize_resource_url(row['url'])
            if normalized_url != row['url']:
                conn.execute('UPDATE resources SET url = ? WHERE id = ?', (normalized_url, row['id']))
        default_rooms = ['Переговорка 1', 'Переговорка 2', 'Конференц-зал', GYM_ROOM_NAME]
        for room_name in default_rooms:
            conn.execute('INSERT OR IGNORE INTO meeting_rooms (name) VALUES (?)', (room_name,))
        driver_resource_exists = conn.execute(
            'SELECT 1 FROM resources WHERE TRIM(url) = ? LIMIT 1',
            ('/driver-trips',)
        ).fetchone()
        if not driver_resource_exists:
            conn.execute(
                'INSERT INTO resources (title, url, category, desc, position) VALUES (?, ?, ?, ?, ?)',
                (
                    'Сервис водителей',
                    '/driver-trips',
                    'Сервисы',
                    'Рейсы и командировки водителей за пределы г. Минска',
                    0
                )
            )
            conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', ('Сервисы',))
        ai_resource_exists = conn.execute(
            'SELECT 1 FROM resources WHERE TRIM(url) = ? LIMIT 1',
            ('/ai-assistant',)
        ).fetchone()
        if not ai_resource_exists:
            conn.execute(
                'INSERT INTO resources (title, url, category, desc, position) VALUES (?, ?, ?, ?, ?)',
                (
                    'БелнипиAI',
                    '/ai-assistant',
                    'Сервисы',
                    'Интеллектуальный помощник по документам и вопросам',
                    0
                )
            )
            conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', ('Сервисы',))
        ensure_default_users_group(conn)
        users_group_id = get_users_group_id(conn)
        if users_group_id:
            members_count = conn.execute(
                'SELECT COUNT(*) AS cnt FROM group_members WHERE group_id = ?',
                (users_group_id,),
            ).fetchone()['cnt']
            if members_count == 0:
                sync_all_ad_users_to_default_group(conn)
        conn.commit()


def get_user_ad_groups(conn, user_dn):
    search_filter = f"(member:1.2.840.113556.1.4.1941:={user_dn})"
    conn.search(LDAP_CONFIG['base'], search_filter, SUBTREE, attributes=['sAMAccountName', 'cn'])
    groups = []
    for entry in conn.entries:
        if entry.sAMAccountName: groups.append(str(entry.sAMAccountName).lower())
        if entry.cn: groups.append(str(entry.cn).lower())
    return list(set(groups))


def get_user_ad_groups_by_username(username):
    login = normalize_ad_username(username)
    if not login:
        return []
    cached = _ad_groups_cache.get(login)
    if isinstance(cached, list):
        return cached
    try:
        server = Server(LDAP_CONFIG['uri'], get_info=ALL, connect_timeout=5)
        conn = Connection(server, user=LDAP_CONFIG['bind_dn'], password=LDAP_CONFIG['bind_password'], auto_bind=True)
        search_filter = f"({LDAP_CONFIG['user_attr']}={login})"
        conn.search(LDAP_CONFIG['base'], search_filter, SUBTREE, attributes=['distinguishedName'])
        if not conn.entries:
            _ad_groups_cache[login] = []
            return []
        user_dn = conn.entries[0].distinguishedName.value
        groups = get_user_ad_groups(conn, user_dn)
        _ad_groups_cache[login] = groups
        return groups
    except Exception:
        return []


def check_ldap_auth(username, password):
    try:
        server = Server(LDAP_CONFIG['uri'], get_info=ALL, connect_timeout=5)
        conn = Connection(server, user=LDAP_CONFIG['bind_dn'], password=LDAP_CONFIG['bind_password'], auto_bind=True)
        search_filter = f"({LDAP_CONFIG['user_attr']}={username})"
        conn.search(LDAP_CONFIG['base'], search_filter, SUBTREE, attributes=['distinguishedName', 'displayName', 'cn'])
        if not conn.entries:
            return False, 'Пользователь не найден в AD'
        user_entry = conn.entries[0]
        user_dn = user_entry.distinguishedName.value
        display_name = ''
        if user_entry.displayName:
            display_name = str(user_entry.displayName).strip()
        if not display_name and user_entry.cn:
            display_name = str(user_entry.cn).strip()
        Connection(server, user=user_dn, password=password, auto_bind=True)
        session['display_name'] = display_name or username
        _ad_groups_cache[normalize_ad_username(username)] = get_user_ad_groups(conn, user_dn)
        session.permanent = True
        return True, ''
    except LDAPException as exc:
        app.logger.warning('LDAP auth failed for "%s": %s', username, exc)
        return False, 'Неверный логин/пароль AD или учетная запись недоступна'
    except Exception as exc:
        app.logger.exception('Unexpected auth error for "%s": %s', username, exc)
        return False, 'Временная ошибка LDAP. Попробуйте позже'


NBRB_RATES_API = 'https://api.nbrb.by/exrates/rates'
NBRB_DISPLAY_CURRENCIES = ('USD', 'EUR', 'RUB')
_nbrb_rates_cache = {}
_nbrb_cache_lock = threading.Lock()
NBRB_CACHE_TTL_SEC = 600


def _parse_nbrb_request_date(raw):
    if not raw:
        return datetime.now(APP_TZ).date(), None
    try:
        return datetime.strptime(raw.strip(), '%Y-%m-%d').date(), None
    except ValueError:
        return None, 'Неверный формат даты'


def _format_nbrb_rate_display(scale, official):
    if scale <= 1:
        text = f'{official:.4f}'.rstrip('0').rstrip('.')
        return text if text else '0'
    text = f'{official:.4f}'.rstrip('0').rstrip('.')
    return f'{text} за {scale}'


def _fetch_nbrb_rates(on_date: date_type):
    cache_key = on_date.isoformat()
    now_ts = time.time()
    with _nbrb_cache_lock:
        cached = _nbrb_rates_cache.get(cache_key)
        if cached and (now_ts - cached['ts']) < NBRB_CACHE_TTL_SEC:
            return cached['data'], None
    url = f'{NBRB_RATES_API}?ondate={cache_key}&periodicity=0'
    try:
        req = urllib_request.Request(
            url,
            headers={'Accept': 'application/json', 'User-Agent': 'BelnipiPortal/1.0'},
        )
        with urllib_request.urlopen(req, timeout=20) as resp:
            payload = json.loads(resp.read().decode('utf-8'))
    except HTTPError as exc:
        return None, f'НБ РБ: HTTP {exc.code}'
    except URLError as exc:
        return None, f'НБ РБ недоступен: {exc.reason}'
    except (TimeoutError, json.JSONDecodeError, OSError, ValueError) as exc:
        return None, str(exc)
    if not isinstance(payload, list):
        return None, 'Неожиданный ответ НБ РБ'
    by_abbr = {
        item.get('Cur_Abbreviation'): item
        for item in payload
        if isinstance(item, dict) and item.get('Cur_Abbreviation')
    }
    rows = []
    for code in NBRB_DISPLAY_CURRENCIES:
        item = by_abbr.get(code)
        if not item:
            continue
        scale = int(item.get('Cur_Scale') or 1)
        official = float(item.get('Cur_OfficialRate') or 0)
        per_one = official / scale if scale else official
        rows.append({
            'code': code,
            'name': (item.get('Cur_Name') or code).strip(),
            'scale': scale,
            'rate': official,
            'rate_per_unit': round(per_one, 4),
            'rate_display': _format_nbrb_rate_display(scale, official),
        })
    data = {
        'date': cache_key,
        'rates': rows,
        'source_url': 'https://www.nbrb.by',
    }
    with _nbrb_cache_lock:
        _nbrb_rates_cache[cache_key] = {'ts': now_ts, 'data': data}
    return data, None


@app.route('/api/nbrb-rates')
def nbrb_rates_api():
    if not session.get('logged_in'):
        return jsonify(success=False, error='Требуется авторизация'), 403
    on_date, err = _parse_nbrb_request_date(request.args.get('date'))
    if err:
        return jsonify(success=False, error=err), 400
    today = datetime.now(APP_TZ).date()
    if on_date > today:
        return jsonify(success=False, error='Дата не может быть в будущем'), 400
    data, fetch_err = _fetch_nbrb_rates(on_date)
    if fetch_err:
        return jsonify(success=False, error=fetch_err), 502
    return jsonify(success=True, **data)


@app.route('/')
def index():
    if not session.get('logged_in'): return redirect(url_for('login_page'))
    username = session.get('username')
    is_admin = username in MASTER_ADMINS
    can_manage_resources_flag = can_manage_resources(username)
    return render_template(
        'index.html',
        is_admin=is_admin,
        can_manage_resources=can_manage_resources_flag,
        user=username
    )


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'GET':
        if session.get('logged_in'): return redirect(url_for('index'))
        return render_template('login.html')
    data = request.json
    u = normalize_ad_username(data.get('username', ''))
    p = data.get('password', '')
    if not u or not p:
        return jsonify(success=False, error='Введите логин и пароль'), 400
    ok, auth_error = check_ldap_auth(u, p)
    if ok:
        session['logged_in'] = True
        session['username'] = u
        conn = get_db_connection()
        try:
            add_user_to_default_group(conn, u)
            conn.commit()
        finally:
            conn.close()
        return jsonify(success=True)
    return jsonify(success=False, error=auth_error or 'Ошибка авторизации AD'), 401


@app.route('/manage')
def manage_page():
    if not session.get('logged_in') or session.get('username') not in MASTER_ADMINS:
        return redirect(url_for('index'))
    return render_template('manage.html', user=session.get('username'))


@app.route('/manage/categories')
def manage_categories_page():
    if not session.get('logged_in') or session.get('username') not in MASTER_ADMINS:
        return redirect(url_for('index'))
    return render_template('manage_categories.html', user=session.get('username'))


@app.route('/phonebook')
def phonebook_page():
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    username = session.get('username')
    if not can_access_portal_path(username, '/phonebook'):
        return redirect(url_for('index'))
    is_admin = username in MASTER_ADMINS
    can_view_private_phones = can_view_extended_phonebook(username)
    contacts = get_phonebook_contacts()
    return render_template(
        'phonebook.html',
        is_admin=is_admin,
        user=username,
        contacts=contacts,
        can_view_private_phones=can_view_private_phones
    )


@app.route('/meeting-rooms')
def meeting_rooms_page():
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    if not can_access_portal_path(session.get('username'), '/meeting-rooms'):
        return redirect(url_for('index'))
    is_admin = session.get('username') in MASTER_ADMINS
    return render_template(
        'meeting_rooms.html',
        is_admin=is_admin,
        user_login=session.get('username'),
        user_display=session.get('display_name') or session.get('username'),
        single_room_mode=False,
        fixed_room_name=''
    )


@app.route('/gym-booking')
def gym_booking_page():
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    if not can_access_portal_path(session.get('username'), '/gym-booking'):
        return redirect(url_for('index'))
    ensure_gym_room_exists()
    is_admin = session.get('username') in MASTER_ADMINS
    return render_template(
        'meeting_rooms.html',
        is_admin=is_admin,
        user_login=session.get('username'),
        user_display=session.get('display_name') or session.get('username'),
        single_room_mode=True,
        fixed_room_name=GYM_ROOM_NAME
    )


@app.route('/driver-trips')
def driver_trips_page():
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    current_user = session.get('username')
    if not can_access_portal_path(current_user, '/driver-trips'):
        return redirect(url_for('index'))
    return render_template(
        'driver_trips.html',
        user_login=current_user,
        user_display=session.get('display_name') or current_user,
        can_manage_all=can_manage_all_bookings(current_user)
    )


@app.route('/knowledge-base')
def knowledge_base_page():
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    if not can_access_portal_path(session.get('username'), '/knowledge-base'):
        return redirect(url_for('index'))
    categories = _knowledge_base_collect_categories()
    return render_template(
        'knowledge_base.html',
        user=session.get('username'),
        categories=categories
    )


@app.route('/ai-assistant')
def ai_assistant_page():
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    username = session.get('username')
    if not can_access_portal_path(username, '/ai-assistant'):
        return redirect(url_for('index'))
    redirect_url = _build_ai_sso_url(username, session.get('display_name') or username)
    return redirect(redirect_url, code=302)


@app.route('/tabel')
def tabel_page():
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    if not can_access_portal_path(session.get('username'), '/tabel'):
        return redirect(url_for('index'))
    ensure_tabel_index(blocking=False)
    now = datetime.now()
    current_period_key = f"{str(now.year)[2:]}_{now.month:02d}"
    with TABEL_INDEX_LOCK:
        all_periods = set()
        for dep_data in TABEL_INDEX.values():
            all_periods.update(dep_data.keys())
        periods = sorted(list(all_periods), reverse=True)
    human_periods = [{"key": p, "label": f"{p.split('_')[1]}.20{p.split('_')[0]}"} for p in periods if '_' in p]
    if not human_periods:
        for offset in range(12):
            month_index = now.month - offset
            year = now.year
            while month_index <= 0:
                month_index += 12
                year -= 1
            yy = str(year)[2:]
            mm = f'{month_index:02d}'
            human_periods.append({"key": f"{yy}_{mm}", "label": f"{mm}.{year}"})
    return render_template(
        'tabel.html',
        user=session.get('username'),
        user_display=session.get('display_name') or session.get('username'),
        periods=human_periods,
        leaders=get_tabel_leaders_data(use_network=False),
        current_period=current_period_key,
    )


@app.route('/logout')
def logout():
    login = normalize_ad_username(session.get('username') or '')
    if login in _ad_groups_cache:
        _ad_groups_cache.pop(login, None)
    session.clear()
    return redirect(url_for('login_page'))


@app.route('/api/tabel/meta')
def tabel_meta():
    if not session.get('logged_in'):
        return jsonify(success=False, error='Требуется авторизация'), 403
    if not can_view_tabel(session.get('username')):
        return jsonify(success=False, error='Нет доступа к табелю'), 403
    ensure_tabel_index(blocking=False)
    now = datetime.now()
    index_summary = _tabel_index_summary()
    source_available = bool(index_summary.get('unique_employees'))
    leaders_source_available = bool(TABEL_LEADERS_CACHE)
    current_period = f'{str(now.year)[2:]}_{now.month:02d}'
    with TABEL_INDEX_LOCK:
        all_periods = set()
        for dep_data in TABEL_INDEX.values():
            all_periods.update(dep_data.keys())
    periods = sorted(all_periods, reverse=True)
    payload = []
    for period in periods:
        if '_' not in period:
            continue
        yy, mm = period.split('_', 1)
        if not (yy.isdigit() and mm.isdigit()):
            continue
        payload.append({'key': period, 'label': f'{mm}.20{yy}'})
    # Если файлов табеля нет/недоступны, отдаем последние 12 месяцев как fallback,
    # чтобы UI оставался рабочим и предсказуемым для пользователя.
    if not payload:
        now = datetime.now()
        for offset in range(12):
            month_index = now.month - offset
            year = now.year
            while month_index <= 0:
                month_index += 12
                year -= 1
            yy = str(year)[2:]
            mm = f'{month_index:02d}'
            payload.append({'key': f'{yy}_{mm}', 'label': f'{mm}.{year}'})
    path_status = tabel_path_status(TABEL_BASE_DIR, TABEL_LEADERS_FILE)
    return jsonify({
        'periods': payload,
        'current_period': current_period,
        'source': {
            'base_dir': TABEL_BASE_DIR,
            'leaders_file': TABEL_LEADERS_FILE,
            'cache_file': os.path.abspath(TABEL_CACHE_FILE),
            'sheet_name': TABEL_SHEET_NAME,
            'base_dir_available': source_available,
            'leaders_file_available': leaders_source_available,
            'indexed_employees': index_summary.get('unique_employees', 0),
            'platform': path_status['platform'],
            'base_dir_exists': path_status['base_dir_exists'],
            'leaders_file_exists': path_status['leaders_file_exists'],
            'setup_hint': path_status['setup_hint'],
            'last_smb': dict(TABEL_LAST_SMB_CONNECT) if TABEL_LAST_SMB_CONNECT else None,
        }
    })


@app.route('/api/tabel/leaders')
def tabel_leaders():
    if not session.get('logged_in'):
        return jsonify(success=False, error='Требуется авторизация'), 403
    if not can_view_tabel(session.get('username')):
        return jsonify(success=False, error='Нет доступа к табелю'), 403
    ensure_tabel_index(blocking=False)
    return jsonify(get_tabel_leaders_data(use_network=True))


@app.route('/api/tabel/search-fio')
def tabel_search_fio():
    if not session.get('logged_in'):
        return jsonify([]), 403
    if not can_view_tabel(session.get('username')):
        return jsonify([]), 403
    ensure_tabel_index(blocking=False)
    query = (request.args.get('q') or '').strip().lower()
    if not query:
        return jsonify([])
    suggestions = set()
    with TABEL_INDEX_LOCK:
        for dep_data in TABEL_INDEX.values():
            for records in dep_data.values():
                for record in records:
                    for employee in record.get('employees', []):
                        fio = employee.get('fio') or ''
                        if query in fio.lower():
                            suggestions.add(fio)
    return jsonify(sorted(list(suggestions))[:15])


@app.route('/api/tabel/status-month')
def tabel_status_month():
    if not session.get('logged_in'):
        return jsonify(success=False, error='Требуется авторизация'), 403
    if not can_view_tabel(session.get('username')):
        return jsonify({'error': 'forbidden'}), 403
    ensure_tabel_index(blocking=False)
    fio = (request.args.get('fio') or '').strip()
    period = (request.args.get('period') or '').strip()
    if not fio or not period or '_' not in period:
        return jsonify({'error': 'invalid'}), 400
    yy, mm = period.split('_', 1)
    try:
        target_year = 2000 + int(yy)
        target_month = int(mm)
        days_in_month = calendar.monthrange(target_year, target_month)[1]
    except Exception:
        return jsonify({'error': 'date'}), 400
    now = datetime.now()
    with TABEL_INDEX_LOCK:
        for dept_name, dep_periods in TABEL_INDEX.items():
            records = dep_periods.get(period) or []
            for record in records:
                match = next((emp for emp in record.get('employees', []) if emp.get('fio') == fio), None)
                if not match:
                    continue
                result = []
                days = match.get('days') or []
                for day_number in range(1, days_in_month + 1):
                    is_future = (
                        target_year > now.year or
                        (target_year == now.year and target_month > now.month) or
                        (target_year == now.year and target_month == now.month and day_number > now.day)
                    )
                    day_idx = day_number - 1
                    value = str(days[day_idx]).strip().upper() if day_idx < len(days) else ''
                    if value in ('', '0', '0.0'):
                        label, color = '—', 'red'
                    elif _tabel_is_work_value(value):
                        label, color = 'На работе', 'green'
                    elif value == 'В':
                        label, color = 'Выходной', 'orange'
                    else:
                        label, color = TABEL_STATUS_MAP.get(value, value), 'red'
                    result.append({'day': day_number, 'label': label, 'color': 'future' if is_future else color})
                return jsonify({'days': result, 'fio': fio, 'department': dept_name})
    return jsonify({'error': 'not_found'}), 404


@app.route('/api/leaders')
def tabel_leaders_compat():
    if not session.get('logged_in'):
        return jsonify([]), 403
    if not can_view_tabel(session.get('username')):
        return jsonify([]), 403
    ensure_tabel_index(blocking=False)
    return jsonify(get_tabel_leaders_data(use_network=True))


@app.route('/search_fio')
def tabel_search_fio_compat():
    if not session.get('logged_in'):
        return jsonify([]), 403
    return tabel_search_fio()


@app.route('/status_month')
def tabel_status_month_compat():
    if not session.get('logged_in'):
        return jsonify({'error': 'unauthorized'}), 403
    return tabel_status_month()


@app.route('/api/knowledge-base/categories')
def knowledge_base_categories():
    if not session.get('logged_in'):
        return jsonify([]), 403
    if not can_access_portal_path(session.get('username'), '/knowledge-base'):
        return jsonify([]), 403
    payload = []
    for category in _knowledge_base_collect_categories():
        payload.append({
            'name': category,
            'files_count': len(_knowledge_base_collect_files(category))
        })
    return jsonify(payload)


@app.route('/api/knowledge-base/files')
def knowledge_base_files():
    if not session.get('logged_in'):
        return jsonify([]), 403
    if not can_access_portal_path(session.get('username'), '/knowledge-base'):
        return jsonify([]), 403
    category = (request.args.get('category') or '').strip()
    if not category:
        return jsonify([]), 400
    return jsonify(_knowledge_base_collect_files(category))


@app.route('/api/knowledge-base/search')
def knowledge_base_search():
    if not session.get('logged_in'):
        return jsonify([]), 403
    if not can_access_portal_path(session.get('username'), '/knowledge-base'):
        return jsonify([]), 403
    query = (request.args.get('q') or '').strip().lower()
    all_items = _knowledge_base_collect_all_files()
    if not query:
        return jsonify(all_items)
    filtered = []
    for item in all_items:
        haystack = f"{item.get('category', '')} {item.get('path', '')} {item.get('name', '')}".lower()
        if query in haystack:
            filtered.append(item)
    return jsonify(filtered)


@app.route('/knowledge-base/view/<path:category_name>/<path:file_path>')
def knowledge_base_view_file(category_name, file_path):
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    if not can_access_portal_path(session.get('username'), '/knowledge-base'):
        return redirect(url_for('index'))
    target_path = _knowledge_base_resolve_file_path(category_name, file_path)
    if not target_path:
        return jsonify(success=False, error='Файл не найден'), 404
    return send_file(target_path, mimetype='application/pdf')


@app.route('/knowledge-base/download/<path:category_name>/<path:file_path>')
def knowledge_base_download_file(category_name, file_path):
    if not session.get('logged_in'):
        return redirect(url_for('login_page'))
    if not can_access_portal_path(session.get('username'), '/knowledge-base'):
        return redirect(url_for('index'))
    target_path = _knowledge_base_resolve_file_path(category_name, file_path)
    if not target_path:
        return jsonify(success=False, error='Файл не найден'), 404
    return send_file(target_path, as_attachment=True, download_name=os.path.basename(target_path))


@app.route('/api/meeting-rooms')
def get_meeting_rooms():
    if not session.get('logged_in'):
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute('SELECT id, name FROM meeting_rooms ORDER BY name COLLATE NOCASE').fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


@app.route('/api/meeting-rooms', methods=['POST'])
def create_meeting_room():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False, error='Только администратор может добавлять переговорки'), 403
    data = request.json or {}
    room_name = (data.get('name') or '').strip()
    if not room_name:
        return jsonify(success=False, error='Укажите название переговорки'), 400
    conn = get_db_connection()
    try:
        exists = conn.execute('SELECT 1 FROM meeting_rooms WHERE name = ?', (room_name,)).fetchone()
        if exists:
            return jsonify(success=False, error='Такая переговорка уже существует'), 409
        conn.execute('INSERT INTO meeting_rooms (name) VALUES (?)', (room_name,))
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/meeting-rooms/<int:room_id>', methods=['DELETE'])
def delete_meeting_room(room_id):
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False, error='Только администратор может удалять переговорки'), 403
    conn = get_db_connection()
    try:
        has_bookings = conn.execute('SELECT 1 FROM meeting_bookings WHERE room_id = ? LIMIT 1', (room_id,)).fetchone()
        if has_bookings:
            return jsonify(success=False, error='Нельзя удалить переговорку: есть существующие брони'), 409
        conn.execute('DELETE FROM meeting_rooms WHERE id = ?', (room_id,))
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


def _meeting_has_conflict(conn, booking_payload, booking_id=None):
    params = [
        booking_payload['room_id'],
        booking_payload['meeting_date'],
        booking_payload['end_time'],
        booking_payload['start_time']
    ]
    query = '''
        SELECT 1
        FROM meeting_bookings
        WHERE room_id = ?
          AND meeting_date = ?
          AND COALESCE(booking_status, 'active') = 'active'
          AND NOT (? <= start_time OR ? >= end_time)
    '''
    if booking_id is not None:
        query += ' AND id <> ?'
        params.append(booking_id)
    query += ' LIMIT 1'
    return conn.execute(query, tuple(params)).fetchone() is not None


def _is_booking_in_past(payload):
    try:
        meeting_start = datetime.strptime(
            f"{payload['meeting_date']} {payload['start_time']}",
            "%Y-%m-%d %H:%M"
        )
    except (TypeError, ValueError):
        return False
    return meeting_start < datetime.now()


def _is_driver_trip_in_past(trip_date, departure_time):
    try:
        trip_start = datetime.strptime(f"{trip_date} {departure_time}", "%Y-%m-%d %H:%M")
    except (TypeError, ValueError):
        return False
    return trip_start < datetime.now()


def _driver_trip_has_conflict(conn, payload, trip_id=None):
    """
    Конфликт для одного авто считаем по фиксированному окну 2 часа от времени выезда.
    """
    try:
        start_dt = datetime.strptime(f"{payload['trip_date']} {payload['departure_time']}", "%Y-%m-%d %H:%M")
    except (TypeError, ValueError):
        return False
    end_dt = start_dt + timedelta(hours=2)
    start_time = start_dt.strftime("%H:%M")
    end_time = end_dt.strftime("%H:%M")
    params = [
        payload['vehicle_model'].strip().lower(),
        payload['trip_date'],
        end_time,
        start_time
    ]
    query = '''
        SELECT 1
        FROM driver_trips
        WHERE LOWER(TRIM(vehicle_model)) = ?
          AND trip_date = ?
          AND COALESCE(trip_status, 'active') = 'active'
          AND NOT (? <= departure_time OR ? >= time(departure_time, '+2 hours'))
    '''
    if trip_id is not None:
        query += ' AND id <> ?'
        params.append(trip_id)
    query += ' LIMIT 1'
    return conn.execute(query, tuple(params)).fetchone() is not None


def _get_booked_by_display_name():
    return (session.get('display_name') or session.get('username') or '').strip()


def _normalize_participants(participants):
    if not isinstance(participants, list):
        return []
    normalized = []
    seen_logins = set()
    for item in participants:
        if not isinstance(item, dict):
            continue
        login = str(item.get('login') or '').strip().lower()
        name = str(item.get('name') or '').strip()
        if not login or not name or login in seen_logins:
            continue
        seen_logins.add(login)
        normalized.append({'login': login, 'name': name})
    return normalized


def _decode_participants(raw_value):
    if not raw_value:
        return []
    try:
        parsed = json.loads(raw_value)
    except (json.JSONDecodeError, TypeError):
        return []
    return _normalize_participants(parsed)


def _serialize_booking_state(booking_row):
    if not booking_row:
        return {}
    item = dict(booking_row)
    item['participants'] = _decode_participants(item.get('participants_json'))
    item.pop('participants_json', None)
    return item


def _log_booking_history(conn, booking_id, action, changed_by, details_payload):
    safe_details = json.dumps(details_payload or {}, ensure_ascii=False)
    conn.execute(
        '''
        INSERT INTO meeting_booking_history (booking_id, action, changed_by, changed_at, details_json)
        VALUES (?, ?, ?, ?, ?)
        ''',
        (
            booking_id,
            action,
            (changed_by or '').strip().lower(),
            datetime.now(APP_TZ).strftime('%Y-%m-%d %H:%M:%S'),
            safe_details
        )
    )


def _serialize_driver_trip_state(trip_row):
    if not trip_row:
        return {}
    return dict(trip_row)


def _log_driver_trip_history(conn, trip_id, action, changed_by, details_payload):
    safe_details = json.dumps(details_payload or {}, ensure_ascii=False)
    conn.execute(
        '''
        INSERT INTO driver_trip_history (trip_id, action, changed_by, changed_at, details_json)
        VALUES (?, ?, ?, ?, ?)
        ''',
        (
            trip_id,
            action,
            (changed_by or '').strip().lower(),
            datetime.now(APP_TZ).strftime('%Y-%m-%d %H:%M:%S'),
            safe_details
        )
    )


def _username_to_corporate_email(username):
    login = (username or '').strip().lower()
    if not login:
        return ''
    return f'{login}@{MAIL_DOMAIN}'


def _send_meeting_cancellation_email(recipient_email, booking_info):
    if not recipient_email:
        return False, 'Не указан email получателя'
    meeting_date = booking_info.get('meeting_date') or ''
    start_time = booking_info.get('start_time') or ''
    end_time = booking_info.get('end_time') or ''
    room_name = booking_info.get('room_name') or ''
    purpose = booking_info.get('purpose') or ''
    canceled_by = booking_info.get('canceled_by') or 'администратор'
    sent_at_minsk = datetime.now(APP_TZ)
    sent_at_text = sent_at_minsk.strftime('%d.%m.%Y %H:%M')
    subject = f'[{sent_at_text} РБ] Отмена встречи: {purpose or "без темы"}'
    body = (
        'Здравствуйте.\n\n'
        'Ваша встреча была отменена администратором.\n\n'
        f'Дата: {meeting_date}\n'
        f'Время: {start_time} - {end_time}\n'
        f'Переговорка: {room_name}\n'
        f'Тема: {purpose}\n'
        f'Кто отменил: {canceled_by}\n'
        f'Время отправки (РБ): {sent_at_text}\n\n'
        'Сообщение отправлено автоматически.'
    )
    message = EmailMessage(policy=policy.SMTPUTF8)
    message['Subject'] = str(Header(subject, 'utf-8'))
    message['From'] = MAIL_SENDER
    message['To'] = recipient_email
    message['Date'] = format_datetime(sent_at_minsk)
    # Принудительно отправляем тело в UTF-8 Base64 для стабильной кириллицы в разных клиентах.
    message.set_content(body, subtype='plain', charset='utf-8', cte='base64')
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as smtp:
            if SMTP_USE_TLS:
                smtp.starttls()
            if SMTP_USERNAME:
                smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
            smtp.send_message(message)
        return True, ''
    except Exception as exc:
        return False, str(exc)


@app.route('/api/meeting-bookings')
def get_meeting_bookings():
    if not session.get('logged_in'):
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute('''
            SELECT
                b.id, b.room_id, r.name AS room_name, b.booked_by, b.purpose,
                b.participants_json, b.meeting_date, b.start_time, b.end_time,
                b.booking_status, b.canceled_by, b.canceled_at, b.owner_username
            FROM meeting_bookings b
            JOIN meeting_rooms r ON r.id = b.room_id
            ORDER BY b.meeting_date ASC, b.start_time ASC
        ''').fetchall()
        data = []
        for row in rows:
            item = dict(row)
            item['participants'] = _decode_participants(item.get('participants_json'))
            item.pop('participants_json', None)
            data.append(item)
        return jsonify(data)
    finally:
        conn.close()


@app.route('/api/meeting-bookings', methods=['POST'])
def create_meeting_booking():
    if not session.get('logged_in'):
        return jsonify(success=False), 403
    data = request.json or {}
    payload = {
        'room_id': data.get('room_id'),
        'purpose': (data.get('purpose') or '').strip(),
        'participants': _normalize_participants(data.get('participants')),
        'meeting_date': (data.get('meeting_date') or '').strip(),
        'start_time': (data.get('start_time') or '').strip(),
        'end_time': (data.get('end_time') or '').strip(),
    }
    if not all([payload['room_id'], payload['purpose'], payload['meeting_date'],
                payload['start_time'], payload['end_time']]):
        return jsonify(success=False, error='Заполните все поля бронирования'), 400
    if not payload['participants']:
        return jsonify(success=False, error='Добавьте хотя бы одного участника из AD'), 400
    booked_by = _get_booked_by_display_name()
    if not booked_by:
        return jsonify(success=False, error='Не удалось определить текущего пользователя AD'), 400
    if payload['end_time'] <= payload['start_time']:
        return jsonify(success=False, error='Время окончания должно быть больше времени начала'), 400
    if _is_booking_in_past(payload):
        return jsonify(success=False, error='Нельзя создавать бронь на прошедшие дату и время'), 400
    conn = get_db_connection()
    try:
        room_exists = conn.execute('SELECT 1 FROM meeting_rooms WHERE id = ?', (payload['room_id'],)).fetchone()
        if not room_exists:
            return jsonify(success=False, error='Выбранная переговорка не найдена'), 404
        if _meeting_has_conflict(conn, payload):
            return jsonify(success=False, error='На выбранное время переговорка уже занята'), 409
        conn.execute('''
            INSERT INTO meeting_bookings (
                room_id, booked_by, purpose, participants_json, meeting_date, start_time, end_time, owner_username
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            payload['room_id'], booked_by, payload['purpose'], json.dumps(payload['participants'], ensure_ascii=False),
            payload['meeting_date'], payload['start_time'], payload['end_time'], session.get('username')
        ))
        booking_id = conn.execute('SELECT last_insert_rowid() AS id').fetchone()['id']
        created_row = conn.execute('''
            SELECT
                b.id, b.room_id, r.name AS room_name, b.booked_by, b.purpose,
                b.participants_json, b.meeting_date, b.start_time, b.end_time,
                b.booking_status, b.canceled_by, b.canceled_at, b.owner_username
            FROM meeting_bookings b
            JOIN meeting_rooms r ON r.id = b.room_id
            WHERE b.id = ?
        ''', (booking_id,)).fetchone()
        _log_booking_history(conn, booking_id, 'created', session.get('username'), {
            'after': _serialize_booking_state(created_row)
        })
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/meeting-bookings/<int:booking_id>', methods=['PUT'])
def update_meeting_booking(booking_id):
    if not session.get('logged_in'):
        return jsonify(success=False), 403
    data = request.json or {}
    payload = {
        'room_id': data.get('room_id'),
        'purpose': (data.get('purpose') or '').strip(),
        'participants': _normalize_participants(data.get('participants')),
        'meeting_date': (data.get('meeting_date') or '').strip(),
        'start_time': (data.get('start_time') or '').strip(),
        'end_time': (data.get('end_time') or '').strip(),
    }
    if not all([payload['room_id'], payload['purpose'], payload['meeting_date'],
                payload['start_time'], payload['end_time']]):
        return jsonify(success=False, error='Заполните все поля бронирования'), 400
    if not payload['participants']:
        return jsonify(success=False, error='Добавьте хотя бы одного участника из AD'), 400
    if payload['end_time'] <= payload['start_time']:
        return jsonify(success=False, error='Время окончания должно быть больше времени начала'), 400
    if _is_booking_in_past(payload):
        return jsonify(success=False, error='Нельзя сохранять бронь на прошедшие дату и время'), 400
    conn = get_db_connection()
    try:
        booking = conn.execute('''
            SELECT
                b.id, b.room_id, r.name AS room_name, b.booked_by, b.purpose,
                b.participants_json, b.meeting_date, b.start_time, b.end_time,
                COALESCE(b.booking_status, 'active') AS booking_status,
                b.canceled_by, b.canceled_at, b.owner_username
            FROM meeting_bookings b
            JOIN meeting_rooms r ON r.id = b.room_id
            WHERE b.id = ?
        ''', (booking_id,)).fetchone()
        if not booking:
            return jsonify(success=False, error='Бронь не найдена'), 404
        if booking['booking_status'] == 'canceled':
            return jsonify(success=False, error='Отмененную бронь редактировать нельзя'), 409
        current_user = session.get('username')
        if not can_manage_all_bookings(current_user) and booking['owner_username'] != current_user:
            return jsonify(success=False, error='Редактировать можно только свои брони'), 403
        room_exists = conn.execute('SELECT 1 FROM meeting_rooms WHERE id = ?', (payload['room_id'],)).fetchone()
        if not room_exists:
            return jsonify(success=False, error='Выбранная переговорка не найдена'), 404
        if _meeting_has_conflict(conn, payload, booking_id=booking_id):
            return jsonify(success=False, error='На выбранное время переговорка уже занята'), 409
        conn.execute('''
            UPDATE meeting_bookings
            SET room_id = ?, purpose = ?, participants_json = ?, meeting_date = ?, start_time = ?, end_time = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            payload['room_id'], payload['purpose'], json.dumps(payload['participants'], ensure_ascii=False),
            payload['meeting_date'], payload['start_time'], payload['end_time'], booking_id
        ))
        updated_row = conn.execute('''
            SELECT
                b.id, b.room_id, r.name AS room_name, b.booked_by, b.purpose,
                b.participants_json, b.meeting_date, b.start_time, b.end_time,
                COALESCE(b.booking_status, 'active') AS booking_status,
                b.canceled_by, b.canceled_at, b.owner_username
            FROM meeting_bookings b
            JOIN meeting_rooms r ON r.id = b.room_id
            WHERE b.id = ?
        ''', (booking_id,)).fetchone()
        _log_booking_history(conn, booking_id, 'updated', current_user, {
            'before': _serialize_booking_state(booking),
            'after': _serialize_booking_state(updated_row)
        })
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/meeting-bookings/<int:booking_id>', methods=['DELETE'])
def delete_meeting_booking(booking_id):
    if not session.get('logged_in'):
        return jsonify(success=False), 403
    conn = get_db_connection()
    try:
        booking = conn.execute('''
            SELECT
                b.id, b.room_id, r.name AS room_name, b.booked_by, b.purpose,
                b.participants_json, b.meeting_date, b.start_time, b.end_time,
                COALESCE(b.booking_status, 'active') AS booking_status,
                b.canceled_by, b.canceled_at, b.owner_username
            FROM meeting_bookings b
            JOIN meeting_rooms r ON r.id = b.room_id
            WHERE b.id = ?
        ''', (booking_id,)).fetchone()
        if not booking:
            return jsonify(success=False, error='Бронь не найдена'), 404
        current_user = session.get('username')
        if not can_manage_all_bookings(current_user) and booking['owner_username'] != current_user:
            return jsonify(success=False, error='Можно отменять только свои брони'), 403
        if booking['booking_status'] == 'canceled':
            return jsonify(success=False, error='Бронь уже отменена'), 409
        conn.execute('''
            UPDATE meeting_bookings
            SET booking_status = 'canceled',
                canceled_by = ?,
                canceled_at = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (current_user, datetime.now(APP_TZ).strftime('%Y-%m-%d %H:%M:%S'), booking_id))
        booking_info = conn.execute('''
            SELECT b.meeting_date, b.start_time, b.end_time, b.purpose,
                   r.name AS room_name, b.owner_username
            FROM meeting_bookings b
            JOIN meeting_rooms r ON r.id = b.room_id
            WHERE b.id = ?
        ''', (booking_id,)).fetchone()
        canceled_row = conn.execute('''
            SELECT
                b.id, b.room_id, r.name AS room_name, b.booked_by, b.purpose,
                b.participants_json, b.meeting_date, b.start_time, b.end_time,
                COALESCE(b.booking_status, 'active') AS booking_status,
                b.canceled_by, b.canceled_at, b.owner_username
            FROM meeting_bookings b
            JOIN meeting_rooms r ON r.id = b.room_id
            WHERE b.id = ?
        ''', (booking_id,)).fetchone()
        _log_booking_history(conn, booking_id, 'canceled', current_user, {
            'before': _serialize_booking_state(booking),
            'after': _serialize_booking_state(canceled_row)
        })
        conn.commit()
        email_warning = ''
        should_notify_owner = (
            booking_info is not None and
            can_manage_all_bookings(current_user) and
            booking_info['owner_username'] != current_user
        )
        if should_notify_owner:
            owner_email = _username_to_corporate_email(booking_info['owner_username'])
            ok, error_text = _send_meeting_cancellation_email(owner_email, {
                'meeting_date': booking_info['meeting_date'],
                'start_time': booking_info['start_time'],
                'end_time': booking_info['end_time'],
                'purpose': booking_info['purpose'],
                'room_name': booking_info['room_name'],
                'canceled_by': current_user
            })
            if not ok:
                email_warning = f'Не удалось отправить уведомление: {error_text}'
        return jsonify(success=True, status='canceled', warning=email_warning)
    finally:
        conn.close()


@app.route('/api/meeting-bookings/<int:booking_id>/history')
def get_meeting_booking_history(booking_id):
    if not session.get('logged_in'):
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        booking_exists = conn.execute('SELECT 1 FROM meeting_bookings WHERE id = ?', (booking_id,)).fetchone()
        if not booking_exists:
            return jsonify([]), 404
        rows = conn.execute('''
            SELECT id, booking_id, action, changed_by, changed_at, details_json
            FROM meeting_booking_history
            WHERE booking_id = ?
            ORDER BY id DESC
        ''', (booking_id,)).fetchall()
        data = []
        for row in rows:
            item = dict(row)
            try:
                item['details'] = json.loads(item.get('details_json') or '{}')
            except (TypeError, json.JSONDecodeError):
                item['details'] = {}
            item.pop('details_json', None)
            data.append(item)
        return jsonify(data)
    finally:
        conn.close()


@app.route('/api/driver-trips')
def get_driver_trips():
    if not session.get('logged_in'):
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            '''
            SELECT id, vehicle_model, vehicle_color, trip_date, departure_time, origin, route_stops, destination, description,
                   created_by, owner_username, trip_status, canceled_by, canceled_at
            FROM driver_trips
            ORDER BY trip_date ASC, departure_time ASC
            '''
        ).fetchall()
        return jsonify([dict(row) for row in rows])
    finally:
        conn.close()


@app.route('/api/driver-trips', methods=['POST'])
def create_driver_trip():
    if not session.get('logged_in'):
        return jsonify(success=False), 403
    data = request.json or {}
    payload = {
        'vehicle_model': (data.get('vehicle_model') or '').strip(),
        'vehicle_color': (data.get('vehicle_color') or '#1f77b4').strip(),
        'trip_date': (data.get('trip_date') or '').strip(),
        'departure_time': (data.get('departure_time') or '').strip(),
        'origin': (data.get('origin') or 'РУП «Белнипиэнергопром»').strip() or 'РУП «Белнипиэнергопром»',
        'route_stops': (data.get('route_stops') or '').strip(),
        'destination': (data.get('destination') or '').strip(),
        'description': (data.get('description') or '').strip()
    }
    if not all([payload['vehicle_model'], payload['trip_date'], payload['departure_time'], payload['destination']]):
        return jsonify(success=False, error='Заполните обязательные поля рейса'), 400
    try:
        datetime.strptime(payload['trip_date'], '%Y-%m-%d')
        datetime.strptime(payload['departure_time'], '%H:%M')
    except ValueError:
        return jsonify(success=False, error='Некорректная дата или время'), 400
    if not re.match(r'^#[0-9a-fA-F]{6}$', payload['vehicle_color']):
        return jsonify(success=False, error='Некорректный цвет автомобиля'), 400
    if _is_driver_trip_in_past(payload['trip_date'], payload['departure_time']):
        return jsonify(success=False, error='Нельзя создавать рейс на прошедшие дату и время'), 400
    conn = get_db_connection()
    try:
        if _driver_trip_has_conflict(conn, payload):
            return jsonify(success=False, error='Для выбранного авто есть пересечение по времени'), 409
        conn.execute(
            '''
            INSERT INTO driver_trips (
                vehicle_model, vehicle_color, trip_date, departure_time, origin, route_stops, destination, description,
                owner_username, created_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                payload['vehicle_model'],
                payload['vehicle_color'],
                payload['trip_date'],
                payload['departure_time'],
                payload['origin'],
                payload['route_stops'],
                payload['destination'],
                payload['description'],
                session.get('username'),
                session.get('username')
            )
        )
        trip_id = conn.execute('SELECT last_insert_rowid() AS id').fetchone()['id']
        created_trip = conn.execute(
            '''
            SELECT id, vehicle_model, vehicle_color, trip_date, departure_time, origin, route_stops, destination, description,
                   created_by, owner_username, trip_status, canceled_by, canceled_at
            FROM driver_trips
            WHERE id = ?
            ''',
            (trip_id,)
        ).fetchone()
        _log_driver_trip_history(conn, trip_id, 'created', session.get('username'), {
            'after': _serialize_driver_trip_state(created_trip)
        })
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/driver-trips/<int:trip_id>', methods=['PUT'])
def update_driver_trip(trip_id):
    if not session.get('logged_in'):
        return jsonify(success=False), 403
    data = request.json or {}
    payload = {
        'vehicle_model': (data.get('vehicle_model') or '').strip(),
        'vehicle_color': (data.get('vehicle_color') or '#1f77b4').strip(),
        'trip_date': (data.get('trip_date') or '').strip(),
        'departure_time': (data.get('departure_time') or '').strip(),
        'origin': (data.get('origin') or 'РУП «Белнипиэнергопром»').strip() or 'РУП «Белнипиэнергопром»',
        'route_stops': (data.get('route_stops') or '').strip(),
        'destination': (data.get('destination') or '').strip(),
        'description': (data.get('description') or '').strip()
    }
    if not all([payload['vehicle_model'], payload['trip_date'], payload['departure_time'], payload['destination']]):
        return jsonify(success=False, error='Заполните обязательные поля рейса'), 400
    try:
        datetime.strptime(payload['trip_date'], '%Y-%m-%d')
        datetime.strptime(payload['departure_time'], '%H:%M')
    except ValueError:
        return jsonify(success=False, error='Некорректная дата или время'), 400
    if not re.match(r'^#[0-9a-fA-F]{6}$', payload['vehicle_color']):
        return jsonify(success=False, error='Некорректный цвет автомобиля'), 400
    if _is_driver_trip_in_past(payload['trip_date'], payload['departure_time']):
        return jsonify(success=False, error='Нельзя редактировать рейс на прошедшие дату и время'), 400
    current_user = session.get('username')
    conn = get_db_connection()
    try:
        trip = conn.execute(
            '''
            SELECT id, vehicle_model, vehicle_color, trip_date, departure_time, origin, route_stops, destination, description,
                   created_by, owner_username, COALESCE(trip_status, 'active') AS trip_status, canceled_by, canceled_at
            FROM driver_trips
            WHERE id = ?
            ''',
            (trip_id,)
        ).fetchone()
        if not trip:
            return jsonify(success=False, error='Рейс не найден'), 404
        owner_login = trip['owner_username'] or trip['created_by']
        if not can_manage_all_bookings(current_user) and owner_login != current_user:
            return jsonify(success=False, error='Редактировать можно только свои рейсы'), 403
        if trip['trip_status'] == 'canceled':
            return jsonify(success=False, error='Отмененный рейс редактировать нельзя'), 409
        if _driver_trip_has_conflict(conn, payload, trip_id=trip_id):
            return jsonify(success=False, error='Для выбранного авто есть пересечение по времени'), 409
        conn.execute(
            '''
            UPDATE driver_trips
            SET vehicle_model = ?, vehicle_color = ?, trip_date = ?, departure_time = ?, origin = ?, route_stops = ?, destination = ?,
                description = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            ''',
            (
                payload['vehicle_model'],
                payload['vehicle_color'],
                payload['trip_date'],
                payload['departure_time'],
                payload['origin'],
                payload['route_stops'],
                payload['destination'],
                payload['description'],
                trip_id
            )
        )
        updated_trip = conn.execute(
            '''
            SELECT id, vehicle_model, vehicle_color, trip_date, departure_time, origin, route_stops, destination, description,
                   created_by, owner_username, COALESCE(trip_status, 'active') AS trip_status, canceled_by, canceled_at
            FROM driver_trips
            WHERE id = ?
            ''',
            (trip_id,)
        ).fetchone()
        _log_driver_trip_history(conn, trip_id, 'updated', current_user, {
            'before': _serialize_driver_trip_state(trip),
            'after': _serialize_driver_trip_state(updated_trip)
        })
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/driver-trips/<int:trip_id>', methods=['DELETE'])
def cancel_driver_trip(trip_id):
    if not session.get('logged_in'):
        return jsonify(success=False), 403
    current_user = session.get('username')
    conn = get_db_connection()
    try:
        trip = conn.execute(
            '''
            SELECT id, vehicle_model, vehicle_color, trip_date, departure_time, origin, route_stops, destination, description,
                   created_by, owner_username, COALESCE(trip_status, 'active') AS trip_status, canceled_by, canceled_at
            FROM driver_trips
            WHERE id = ?
            ''',
            (trip_id,)
        ).fetchone()
        if not trip:
            return jsonify(success=False, error='Рейс не найден'), 404
        owner_login = trip['owner_username'] or trip['created_by']
        if not can_manage_all_bookings(current_user) and owner_login != current_user:
            return jsonify(success=False, error='Можно отменять только свои рейсы'), 403
        if trip['trip_status'] == 'canceled':
            return jsonify(success=False, error='Рейс уже отменен'), 409
        conn.execute(
            '''
            UPDATE driver_trips
            SET trip_status = 'canceled',
                canceled_by = ?,
                canceled_at = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            ''',
            (current_user, datetime.now(APP_TZ).strftime('%Y-%m-%d %H:%M:%S'), trip_id)
        )
        canceled_trip = conn.execute(
            '''
            SELECT id, vehicle_model, vehicle_color, trip_date, departure_time, origin, route_stops, destination, description,
                   created_by, owner_username, COALESCE(trip_status, 'active') AS trip_status, canceled_by, canceled_at
            FROM driver_trips
            WHERE id = ?
            ''',
            (trip_id,)
        ).fetchone()
        _log_driver_trip_history(conn, trip_id, 'canceled', current_user, {
            'before': _serialize_driver_trip_state(trip),
            'after': _serialize_driver_trip_state(canceled_trip)
        })
        conn.commit()
        return jsonify(success=True, status='canceled')
    finally:
        conn.close()


@app.route('/api/driver-trips/<int:trip_id>/history')
def get_driver_trip_history(trip_id):
    if not session.get('logged_in'):
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        trip_exists = conn.execute('SELECT 1 FROM driver_trips WHERE id = ?', (trip_id,)).fetchone()
        if not trip_exists:
            return jsonify([]), 404
        rows = conn.execute(
            '''
            SELECT id, trip_id, action, changed_by, changed_at, details_json
            FROM driver_trip_history
            WHERE trip_id = ?
            ORDER BY id DESC
            ''',
            (trip_id,)
        ).fetchall()
        data = []
        for row in rows:
            item = dict(row)
            try:
                item['details'] = json.loads(item.get('details_json') or '{}')
            except (TypeError, json.JSONDecodeError):
                item['details'] = {}
            item.pop('details_json', None)
            data.append(item)
        return jsonify(data)
    finally:
        conn.close()


@app.route('/assets/logo')
def project_logo():
    return send_from_directory(app.root_path, LOGO_FILENAME)


@app.route('/assets/<path:filename>')
def project_asset(filename):
    return send_from_directory(app.root_path, filename)


@app.route('/get_groups')
def get_groups():
    if not session.get('logged_in'): return jsonify([])
    conn = get_db_connection()
    try:
        groups = conn.execute('SELECT * FROM groups').fetchall()
        res = []
        for g in groups:
            m = conn.execute('SELECT username FROM group_members WHERE group_id = ?', (g['id'],)).fetchall()
            res.append({'id': g['id'], 'name': g['name'], 'members': [x['username'] for x in m]})
        return jsonify(res)
    finally:
        conn.close()


@app.route('/get_categories')
def get_categories():
    if not session.get('logged_in'):
        return jsonify([])
    conn = get_db_connection()
    try:
        rows = conn.execute('SELECT name FROM categories ORDER BY name COLLATE NOCASE').fetchall()
        return jsonify([r['name'] for r in rows])
    finally:
        conn.close()


@app.route('/get_categories_overview')
def get_categories_overview():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute('''
            SELECT c.name AS name, COUNT(r.id) AS resource_count
            FROM categories c
            LEFT JOIN resources r ON TRIM(r.category) = c.name
            GROUP BY c.name
            ORDER BY c.name COLLATE NOCASE
        ''').fetchall()
        return jsonify([{'name': r['name'], 'resource_count': r['resource_count']} for r in rows])
    finally:
        conn.close()


@app.route('/get_ad_entities')
def get_ad_entities():
    if not session.get('logged_in'): return jsonify([])
    q = request.args.get('q', '').strip()
    if len(q) < 2: return jsonify([])
    try:
        server = Server(LDAP_CONFIG['uri'], get_info=ALL)
        conn = Connection(server, user=LDAP_CONFIG['bind_dn'], password=LDAP_CONFIG['bind_password'], auto_bind=True)
        search_filter = f"(|(sAMAccountName=*{q}*)(displayName=*{q}*)(cn=*{q}*))"
        conn.search(LDAP_CONFIG['base'], search_filter, SUBTREE,
                    attributes=['sAMAccountName', 'displayName', 'objectClass', 'cn'])
        results = []
        for entry in conn.entries:
            is_group = 'group' in entry.objectClass
            login_val = str(entry.cn) if is_group else str(entry.sAMAccountName)
            results.append({'login': login_val, 'name': str(entry.displayName) if entry.displayName else str(entry.cn),
                            'type': 'Группа AD' if is_group else 'Юзер'})
        return jsonify(results[:20])
    except:
        return jsonify([])


@app.route('/get_resources')
def get_resources():
    if not session.get('logged_in'): return jsonify([]), 403
    u = session.get('username')
    search_query = request.args.get('search', '').strip().lower()
    user_ad_groups = get_user_ad_groups_by_username(u)
    conn = get_db_connection()
    try:
        rows = conn.execute('''
            SELECT r.*, GROUP_CONCAT(ga.group_id) as group_ids
            FROM resources r
            LEFT JOIN resource_group_access ga ON r.id = ga.resource_id
            GROUP BY r.id
            ORDER BY r.position ASC
        ''').fetchall()
        def matches_search(resource_row):
            if not search_query:
                return True
            haystack = " ".join([
                str(resource_row.get('title') or ''),
                str(resource_row.get('desc') or ''),
                str(resource_row.get('category') or ''),
                str(resource_row.get('url') or '')
            ]).lower()
            return search_query in haystack

        if u in MASTER_ADMINS:
            admin_rows = [dict(row) for row in rows]
            for row in admin_rows:
                row['url'] = normalize_resource_url(row.get('url'))
            return jsonify([row for row in admin_rows if matches_search(row)])

        group_map, users_group_id = _load_resource_group_access_maps(conn)

        visible = []
        for r in rows:
            row_dict = dict(r)
            row_dict['url'] = normalize_resource_url(row_dict.get('url'))
            if can_access_portal_resource(
                u,
                row_dict['url'],
                r['group_ids'],
                group_map,
                users_group_id,
                user_ad_groups,
            ) and matches_search(row_dict):
                visible.append(row_dict)
        return jsonify(visible)
    finally:
        conn.close()


@app.route('/add', methods=['POST'])
def add_resource():
    if not can_manage_resources(session.get('username')): return jsonify(success=False), 403
    t = request.form.get('title')
    u = normalize_resource_url(request.form.get('url'))
    c_existing = (request.form.get('category_existing') or '').strip()
    if c_existing == '__new__':
        c_existing = ''
    c_new = (request.form.get('category_new') or '').strip()
    c = c_new or c_existing or (request.form.get('category') or '').strip()
    d = request.form.get('desc')
    if not c:
        return jsonify(success=False, error='Укажите раздел'), 400
    gids = request.form.getlist('access_group_ids')
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO resources (title, url, category, desc) VALUES (?, ?, ?, ?)", (t, u, c, d))
        cur.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (c,))
        rid = cur.lastrowid
        for gid in gids: cur.execute("INSERT INTO resource_group_access (resource_id, group_id) VALUES (?, ?)",
                                     (rid, int(gid)))
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/edit/<int:res_id>', methods=['POST'])
def edit_resource(res_id):
    if not can_manage_resources(session.get('username')): return jsonify(success=False), 403
    t = request.form.get('title')
    u = normalize_resource_url(request.form.get('url'))
    c_existing = (request.form.get('category_existing') or '').strip()
    if c_existing == '__new__':
        c_existing = ''
    c_new = (request.form.get('category_new') or '').strip()
    c = c_new or c_existing or (request.form.get('category') or '').strip()
    d = request.form.get('desc')
    if not c:
        return jsonify(success=False, error='Укажите раздел'), 400
    gids = request.form.getlist('access_group_ids')
    conn = get_db_connection()
    try:
        conn.execute("UPDATE resources SET title=?, url=?, category=?, desc=? WHERE id=?", (t, u, c, d, res_id))
        conn.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (c,))
        conn.execute("DELETE FROM resource_group_access WHERE resource_id=?", (res_id,))
        for gid in gids: conn.execute("INSERT INTO resource_group_access (resource_id, group_id) VALUES (?, ?)",
                                      (res_id, int(gid)))
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/delete/<int:res_id>', methods=['POST'])
def delete_resource(res_id):
    if not can_manage_resources(session.get('username')): return jsonify(success=False), 403
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM resources WHERE id=?", (res_id,))
        conn.execute("DELETE FROM resource_group_access WHERE resource_id=?", (res_id,))
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/reorder', methods=['POST'])
def reorder():
    if not can_manage_resources(session.get('username')): return jsonify(success=False), 403
    conn = get_db_connection()
    try:
        for index, entry in enumerate(request.json):
            conn.execute("UPDATE resources SET position=?, category=? WHERE id=?",
                         (index, entry['category'], entry['id']))
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/manage_category', methods=['POST'])
def manage_category():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    data = request.json or {}
    action = (data.get('action') or '').strip()
    conn = get_db_connection()
    try:
        if action == 'create':
            category_name = (data.get('category_name') or '').strip()
            if not category_name:
                return jsonify(success=False, error='Укажите название нового раздела'), 400
            exists = conn.execute('SELECT 1 FROM categories WHERE name = ?', (category_name,)).fetchone()
            if exists:
                return jsonify(success=False, error='Раздел с таким названием уже существует'), 409
            conn.execute('INSERT INTO categories (name) VALUES (?)', (category_name,))
            conn.commit()
            return jsonify(success=True)

        if action == 'rename':
            old_name = (data.get('old_name') or '').strip()
            new_name = (data.get('new_name') or '').strip()
            if not old_name or not new_name:
                return jsonify(success=False, error='Укажите старое и новое название раздела'), 400
            if old_name == new_name:
                return jsonify(success=False, error='Название раздела не изменилось'), 400
            exists = conn.execute('SELECT 1 FROM categories WHERE name = ?', (old_name,)).fetchone()
            if not exists:
                return jsonify(success=False, error='Раздел не найден'), 404
            conflict = conn.execute('SELECT 1 FROM categories WHERE name = ?', (new_name,)).fetchone()
            if conflict:
                return jsonify(success=False, error='Раздел с таким названием уже существует'), 409
            conn.execute('UPDATE resources SET category = ? WHERE category = ?', (new_name, old_name))
            conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (new_name,))
            conn.execute('DELETE FROM categories WHERE name = ?', (old_name,))
            conn.commit()
            return jsonify(success=True)

        if action == 'delete':
            category_name = (data.get('category_name') or '').strip()
            transfer_mode = (data.get('transfer_mode') or '').strip()
            target_category = (data.get('target_category') or '').strip()
            resource_moves = data.get('resource_moves') or {}
            if not category_name:
                return jsonify(success=False, error='Укажите раздел для удаления'), 400
            exists = conn.execute('SELECT 1 FROM categories WHERE name = ?', (category_name,)).fetchone()
            if not exists:
                return jsonify(success=False, error='Раздел не найден'), 404
            resources = conn.execute(
                'SELECT id FROM resources WHERE category = ? ORDER BY id',
                (category_name,)
            ).fetchall()
            resource_ids = [str(r['id']) for r in resources]

            if resource_ids:
                if transfer_mode == 'single':
                    if not target_category:
                        return jsonify(success=False, error='Выберите целевой раздел для ресурсов'), 400
                    if target_category == category_name:
                        return jsonify(success=False, error='Нельзя переносить в удаляемый раздел'), 400
                    conn.execute('UPDATE resources SET category = ? WHERE category = ?', (target_category, category_name))
                    conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (target_category,))
                elif transfer_mode == 'split':
                    if not isinstance(resource_moves, dict):
                        return jsonify(success=False, error='Некорректные данные распределения'), 400
                    targets_to_create = set()
                    for rid in resource_ids:
                        target = str(resource_moves.get(rid, '')).strip()
                        if not target:
                            return jsonify(success=False, error='Укажите целевой раздел для каждого ресурса'), 400
                        if target == category_name:
                            return jsonify(success=False, error='Нельзя переносить в удаляемый раздел'), 400
                        targets_to_create.add(target)
                    for t in targets_to_create:
                        conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (t,))
                    for rid in resource_ids:
                        target = str(resource_moves.get(rid, '')).strip()
                        conn.execute('UPDATE resources SET category = ? WHERE id = ?', (target, int(rid)))
                elif transfer_mode == 'delete_all':
                    conn.execute(
                        'DELETE FROM resource_group_access WHERE resource_id IN (SELECT id FROM resources WHERE category = ?)',
                        (category_name,)
                    )
                    conn.execute('DELETE FROM resources WHERE category = ?', (category_name,))
                else:
                    return jsonify(success=False, error='Выберите режим обработки ресурсов'), 400

            conn.execute('DELETE FROM categories WHERE name = ?', (category_name,))
            conn.commit()
            return jsonify(success=True)

        return jsonify(success=False, error='Неизвестное действие'), 400
    finally:
        conn.close()


@app.route('/manage_group', methods=['POST'])
def manage_group():
    if session.get('username') not in MASTER_ADMINS: return jsonify(success=False), 403
    data = request.json
    action, name, gid = data.get('action'), data.get('name'), data.get('id')
    conn = get_db_connection()
    try:
        if action == 'add':
            conn.execute('INSERT OR IGNORE INTO groups (name) VALUES (?)', (name,))
        elif action == 'delete':
            conn.execute('DELETE FROM groups WHERE id = ?', (gid,))
            conn.execute('DELETE FROM group_members WHERE group_id = ?', (gid,))
            conn.execute('DELETE FROM resource_group_access WHERE group_id = ?', (gid,))
        elif action == 'update_members':
            conn.execute('DELETE FROM group_members WHERE group_id = ?', (gid,))
            for m in data.get('members', []):
                if m.strip(): conn.execute('INSERT INTO group_members (group_id, username) VALUES (?, ?)',
                                           (gid, m.strip().lower()))
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/phonebook-access')
def get_phonebook_access_users():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            '''
            SELECT entity_type, entity_login
            FROM phonebook_privileged_entities
            ORDER BY entity_type ASC, entity_login COLLATE NOCASE
            '''
        ).fetchall()
        payload = [
            {'type': row['entity_type'], 'login': row['entity_login']}
            for row in rows
        ]
        return jsonify(payload)
    finally:
        conn.close()


@app.route('/manage_phonebook_access', methods=['POST'])
def manage_phonebook_access():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    payload = request.json or {}
    action = (payload.get('action') or '').strip()
    entity_type = (payload.get('type') or 'user').strip().lower()
    raw_login = payload.get('username')
    if entity_type == 'group':
        entity_login = str(raw_login or '').strip().lower()
    else:
        entity_type = 'user'
        entity_login = normalize_ad_username(raw_login)
    if action not in ('add', 'delete'):
        return jsonify(success=False, error='Неизвестное действие'), 400
    if not entity_login:
        return jsonify(success=False, error='Укажите пользователя'), 400
    conn = get_db_connection()
    try:
        if action == 'add':
            conn.execute(
                'INSERT OR IGNORE INTO phonebook_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (entity_type, entity_login)
            )
        else:
            conn.execute(
                'DELETE FROM phonebook_privileged_entities WHERE entity_type = ? AND entity_login = ?',
                (entity_type, entity_login)
            )
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/booking-access')
def get_booking_access_entities():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            '''
            SELECT entity_type, entity_login
            FROM booking_privileged_entities
            ORDER BY entity_type ASC, entity_login COLLATE NOCASE
            '''
        ).fetchall()
        return jsonify([{'type': row['entity_type'], 'login': row['entity_login']} for row in rows])
    finally:
        conn.close()


@app.route('/api/tabel-access')
def get_tabel_access_entities():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            '''
            SELECT entity_type, entity_login
            FROM tabel_privileged_entities
            ORDER BY entity_type ASC, entity_login COLLATE NOCASE
            '''
        ).fetchall()
        return jsonify([{'type': row['entity_type'], 'login': row['entity_login']} for row in rows])
    finally:
        conn.close()


@app.route('/manage_tabel_access', methods=['POST'])
def manage_tabel_access():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    payload = request.json or {}
    action = (payload.get('action') or '').strip()
    entity_type = (payload.get('type') or 'user').strip().lower()
    raw_login = payload.get('username')
    if entity_type == 'group':
        entity_login = str(raw_login or '').strip().lower()
    else:
        entity_type = 'user'
        entity_login = normalize_ad_username(raw_login)
    if action not in ('add', 'delete'):
        return jsonify(success=False, error='Неизвестное действие'), 400
    if not entity_login:
        return jsonify(success=False, error='Укажите пользователя или группу'), 400
    conn = get_db_connection()
    try:
        if action == 'add':
            conn.execute(
                'INSERT OR IGNORE INTO tabel_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (entity_type, entity_login)
            )
        else:
            conn.execute(
                'DELETE FROM tabel_privileged_entities WHERE entity_type = ? AND entity_login = ?',
                (entity_type, entity_login)
            )
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/manage_booking_access', methods=['POST'])
def manage_booking_access():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    payload = request.json or {}
    action = (payload.get('action') or '').strip()
    entity_type = (payload.get('type') or 'user').strip().lower()
    raw_login = payload.get('username')
    if entity_type == 'group':
        entity_login = str(raw_login or '').strip().lower()
    else:
        entity_type = 'user'
        entity_login = normalize_ad_username(raw_login)
    if action not in ('add', 'delete'):
        return jsonify(success=False, error='Неизвестное действие'), 400
    if not entity_login:
        return jsonify(success=False, error='Укажите пользователя или группу'), 400
    conn = get_db_connection()
    try:
        if action == 'add':
            conn.execute(
                'INSERT OR IGNORE INTO booking_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (entity_type, entity_login)
            )
        else:
            conn.execute(
                'DELETE FROM booking_privileged_entities WHERE entity_type = ? AND entity_login = ?',
                (entity_type, entity_login)
            )
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/resource-access')
def get_resource_access_entities():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            '''
            SELECT entity_type, entity_login
            FROM resource_privileged_entities
            ORDER BY entity_type ASC, entity_login COLLATE NOCASE
            '''
        ).fetchall()
        return jsonify([{'type': row['entity_type'], 'login': row['entity_login']} for row in rows])
    finally:
        conn.close()


@app.route('/manage_resource_access', methods=['POST'])
def manage_resource_access():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    payload = request.json or {}
    action = (payload.get('action') or '').strip()
    entity_type = (payload.get('type') or 'user').strip().lower()
    raw_login = payload.get('username')
    if entity_type == 'group':
        entity_login = str(raw_login or '').strip().lower()
    else:
        entity_type = 'user'
        entity_login = normalize_ad_username(raw_login)
    if action not in ('add', 'delete'):
        return jsonify(success=False, error='Неизвестное действие'), 400
    if not entity_login:
        return jsonify(success=False, error='Укажите пользователя или группу'), 400
    conn = get_db_connection()
    try:
        if action == 'add':
            conn.execute(
                'INSERT OR IGNORE INTO resource_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (entity_type, entity_login)
            )
        else:
            conn.execute(
                'DELETE FROM resource_privileged_entities WHERE entity_type = ? AND entity_login = ?',
                (entity_type, entity_login)
            )
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


@app.route('/api/ai-access')
def get_ai_access_entities():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify([]), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            '''
            SELECT entity_type, entity_login
            FROM ai_privileged_entities
            ORDER BY entity_type ASC, entity_login COLLATE NOCASE
            '''
        ).fetchall()
        return jsonify([{'type': row['entity_type'], 'login': row['entity_login']} for row in rows])
    finally:
        conn.close()


ACCESS_EXPORT_TYPE = 'belnipi_access'
ACCESS_EXPORT_VERSION = 1


def _build_access_export_payload(conn):
    groups_rows = conn.execute('SELECT id, name FROM groups ORDER BY name COLLATE NOCASE').fetchall()
    groups_out = [{'name': row['name']} for row in groups_rows]

    members_rows = conn.execute(
        '''
        SELECT gm.username, g.name AS group_name
        FROM group_members gm
        JOIN groups g ON g.id = gm.group_id
        ORDER BY g.name COLLATE NOCASE, gm.username COLLATE NOCASE
        '''
    ).fetchall()
    group_members_out = [
        {'group_name': row['group_name'], 'username': row['username']} for row in members_rows
    ]

    cat_rows = conn.execute('SELECT name FROM categories ORDER BY name COLLATE NOCASE').fetchall()
    categories_out = [{'name': row['name']} for row in cat_rows]

    res_rows = conn.execute(
        '''
        SELECT r.id, r.title, r.url, r.category, r.desc, r.position,
               GROUP_CONCAT(g.name) AS group_names
        FROM resources r
        LEFT JOIN resource_group_access ga ON ga.resource_id = r.id
        LEFT JOIN groups g ON g.id = ga.group_id
        GROUP BY r.id
        ORDER BY r.position ASC, r.id ASC
        '''
    ).fetchall()
    resources_out = []
    for row in res_rows:
        raw_names = row['group_names'] or ''
        gnames = [x.strip() for x in raw_names.split(',') if x and x.strip()]
        resources_out.append({
            'title': row['title'],
            'url': normalize_resource_url(row['url'] or ''),
            'category': (row['category'] or '').strip(),
            'desc': row['desc'],
            'position': row['position'] if row['position'] is not None else 0,
            'groups': gnames,
        })

    def privileged(table):
        rows = conn.execute(
            f'SELECT entity_type, entity_login FROM {table} ORDER BY entity_type, entity_login COLLATE NOCASE'
        ).fetchall()
        return [{'type': r['entity_type'], 'login': r['entity_login']} for r in rows]

    exported_at = datetime.now(APP_TZ).isoformat(timespec='seconds')
    return {
        'export_type': ACCESS_EXPORT_TYPE,
        'version': ACCESS_EXPORT_VERSION,
        'exported_at': exported_at,
        'groups': groups_out,
        'group_members': group_members_out,
        'categories': categories_out,
        'resources': resources_out,
        'phonebook_privileged_entities': privileged('phonebook_privileged_entities'),
        'booking_privileged_entities': privileged('booking_privileged_entities'),
        'resource_privileged_entities': privileged('resource_privileged_entities'),
        'ai_privileged_entities': privileged('ai_privileged_entities'),
        'tabel_privileged_entities': privileged('tabel_privileged_entities'),
    }


@app.route('/api/admin/export-access', methods=['GET'])
def export_access_permissions():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    conn = get_db_connection()
    try:
        payload = _build_access_export_payload(conn)
    finally:
        conn.close()
    body = json.dumps(payload, ensure_ascii=False, indent=2)
    fname = datetime.now(APP_TZ).strftime('access_%Y-%m-%d_%H-%M-%S.json')
    resp = Response(
        body + '\n',
        mimetype='application/json; charset=utf-8',
    )
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@app.route('/api/admin/import-access', methods=['POST'])
def import_access_permissions():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    raw = request.get_data(as_text=True) or ''
    if not raw.strip():
        return jsonify(success=False, error='Пустой файл'), 400
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return jsonify(success=False, error='Некорректный JSON'), 400
    if data.get('export_type') != ACCESS_EXPORT_TYPE:
        return jsonify(success=False, error='Неизвестный тип файла (ожидается belnipi_access)'), 400
    if int(data.get('version') or 0) != ACCESS_EXPORT_VERSION:
        return jsonify(success=False, error='Неподдерживаемая версия экспорта'), 400

    def as_obj_list(key, name_key='name'):
        items = data.get(key) or []
        if not isinstance(items, list):
            return []
        out = []
        for item in items:
            if isinstance(item, str):
                name = item.strip()
                if name:
                    out.append({name_key: name})
            elif isinstance(item, dict):
                name = (item.get(name_key) or item.get('name') or '').strip()
                if name:
                    copy = dict(item)
                    copy[name_key] = name
                    out.append(copy)
        return out

    def as_member_list():
        items = data.get('group_members') or []
        if not isinstance(items, list):
            return []
        out = []
        for item in items:
            if not isinstance(item, dict):
                continue
            gn = (item.get('group_name') or '').strip()
            un = (item.get('username') or '').strip()
            if gn and un:
                out.append({'group_name': gn, 'username': un})
        return out

    def as_privileged(key):
        items = data.get(key) or []
        if not isinstance(items, list):
            return []
        out = []
        for item in items:
            if not isinstance(item, dict):
                continue
            et = (item.get('type') or 'user').strip().lower()
            if et not in ('user', 'group'):
                et = 'user'
            el = (item.get('login') or '').strip()
            if not el:
                continue
            if et == 'user':
                el = normalize_ad_username(el)
            else:
                el = el.lower()
            if el:
                out.append((et, el))
        return out

    def as_resources():
        items = data.get('resources') or []
        if not isinstance(items, list):
            return []
        out = []
        for item in items:
            if not isinstance(item, dict):
                continue
            title = item.get('title') or ''
            url = normalize_resource_url((item.get('url') or '').strip())
            category = (item.get('category') or '').strip()
            if not category:
                continue
            desc = item.get('desc')
            try:
                position = int(item.get('position') or 0)
            except (TypeError, ValueError):
                position = 0
            gnames = item.get('groups') or []
            if isinstance(gnames, str):
                gnames = [gnames]
            if not isinstance(gnames, list):
                gnames = []
            gnames = [str(x).strip() for x in gnames if str(x).strip()]
            out.append({
                'title': title,
                'url': url,
                'category': category,
                'desc': desc,
                'position': position,
                'groups': gnames,
            })
        return out

    groups_in = as_obj_list('groups', 'name')
    members_in = as_member_list()
    categories_in = as_obj_list('categories', 'name')
    resources_in = as_resources()
    phone_in = as_privileged('phonebook_privileged_entities')
    booking_in = as_privileged('booking_privileged_entities')
    resource_in = as_privileged('resource_privileged_entities')
    ai_in = as_privileged('ai_privileged_entities')
    tabel_in = as_privileged('tabel_privileged_entities')

    conn = get_db_connection()
    try:
        conn.execute('BEGIN')
        conn.execute('DELETE FROM resource_group_access')
        conn.execute('DELETE FROM group_members')
        conn.execute('DELETE FROM phonebook_privileged_entities')
        conn.execute('DELETE FROM booking_privileged_entities')
        conn.execute('DELETE FROM resource_privileged_entities')
        conn.execute('DELETE FROM ai_privileged_entities')
        conn.execute('DELETE FROM tabel_privileged_entities')

        for row in groups_in:
            conn.execute('INSERT OR IGNORE INTO groups (name) VALUES (?)', (row['name'],))

        for row in categories_in:
            conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (row['name'],))

        for row in members_in:
            gid = conn.execute('SELECT id FROM groups WHERE name = ?', (row['group_name'],)).fetchone()
            if not gid:
                conn.execute('INSERT OR IGNORE INTO groups (name) VALUES (?)', (row['group_name'],))
                gid = conn.execute('SELECT id FROM groups WHERE name = ?', (row['group_name'],)).fetchone()
            if gid:
                conn.execute(
                    'INSERT OR IGNORE INTO group_members (group_id, username) VALUES (?, ?)',
                    (gid['id'], row['username']),
                )

        for res in resources_in:
            row = conn.execute(
                '''
                SELECT id FROM resources
                WHERE TRIM(url) = TRIM(?) AND TRIM(category) = TRIM(?)
                ORDER BY id ASC LIMIT 1
                ''',
                (res['url'], res['category']),
            ).fetchone()
            if row:
                rid = row['id']
                conn.execute(
                    'UPDATE resources SET title = ?, url = ?, category = ?, desc = ?, position = ? WHERE id = ?',
                    (res['title'], res['url'], res['category'], res['desc'], res['position'], rid),
                )
            else:
                cur = conn.execute(
                    'INSERT INTO resources (title, url, category, desc, position) VALUES (?, ?, ?, ?, ?)',
                    (res['title'], res['url'], res['category'], res['desc'], res['position']),
                )
                rid = cur.lastrowid
            conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (res['category'],))
            for gname in res['groups']:
                gid = conn.execute('SELECT id FROM groups WHERE name = ?', (gname,)).fetchone()
                if not gid:
                    conn.execute('INSERT OR IGNORE INTO groups (name) VALUES (?)', (gname,))
                    gid = conn.execute('SELECT id FROM groups WHERE name = ?', (gname,)).fetchone()
                if gid:
                    conn.execute(
                        'INSERT OR IGNORE INTO resource_group_access (resource_id, group_id) VALUES (?, ?)',
                        (rid, gid['id']),
                    )

        for et, el in phone_in:
            conn.execute(
                'INSERT OR IGNORE INTO phonebook_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (et, el),
            )
        for et, el in booking_in:
            conn.execute(
                'INSERT OR IGNORE INTO booking_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (et, el),
            )
        for et, el in resource_in:
            conn.execute(
                'INSERT OR IGNORE INTO resource_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (et, el),
            )
        for et, el in ai_in:
            conn.execute(
                'INSERT OR IGNORE INTO ai_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (et, el),
            )
        for et, el in tabel_in:
            conn.execute(
                'INSERT OR IGNORE INTO tabel_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (et, el),
            )

        conn.commit()
    except Exception as exc:
        conn.rollback()
        return jsonify(success=False, error=str(exc)), 400
    finally:
        conn.close()

    return jsonify(
        success=True,
        imported={
            'groups': len(groups_in),
            'group_members': len(members_in),
            'categories': len(categories_in),
            'resources': len(resources_in),
            'phonebook_privileged': len(phone_in),
            'booking_privileged': len(booking_in),
            'resource_privileged': len(resource_in),
            'ai_privileged': len(ai_in),
            'tabel_privileged': len(tabel_in),
        },
    )


def _validate_portal_sqlite_file(path):
    """Проверка целостности и минимальной схемы перед заменой рабочей БД."""
    try:
        conn = sqlite3.connect(path, timeout=30)
    except sqlite3.Error as exc:
        return False, f'Не удалось открыть файл как SQLite: {exc}'
    try:
        row = conn.execute('PRAGMA integrity_check').fetchone()
        if not row or row[0] != 'ok':
            return False, (row[0] if row else 'Ошибка проверки целостности')
        n = conn.execute(
            "SELECT count(*) FROM sqlite_master WHERE type='table' "
            "AND name IN ('resources','groups','meeting_rooms')"
        ).fetchone()[0]
        if int(n) < 3:
            return False, 'Файл не похож на базу этого приложения (нет ключевых таблиц).'
        return True, None
    finally:
        conn.close()


def _build_database_export_bytes():
    """Онлайн-копия SQLite через backup API (без обрыва записей в WAL)."""
    abs_path = os.path.abspath(DB_PATH)
    if not os.path.isfile(abs_path):
        return None, 'Файл базы не найден на сервере'
    fd, tmp_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    src = sqlite3.connect(abs_path, timeout=30)
    try:
        dst = sqlite3.connect(tmp_path)
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()
    try:
        with open(tmp_path, 'rb') as f:
            data = f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    return data, None


def _remove_db_wal_shm(db_path):
    for suf in ('-wal', '-shm'):
        p = db_path + suf
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass


def _wal_checkpoint_truncate_and_remove_sidecars(db_path):
    if not os.path.isfile(db_path):
        return
    try:
        conn = sqlite3.connect(db_path, timeout=30)
    except sqlite3.Error:
        return
    try:
        try:
            conn.execute('PRAGMA wal_checkpoint(TRUNCATE)')
        except sqlite3.Error:
            pass
        conn.commit()
    finally:
        conn.close()
    _remove_db_wal_shm(db_path)


@app.route('/api/admin/export-database', methods=['GET'])
def export_database_file():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    data, err = _build_database_export_bytes()
    if err:
        return jsonify(success=False, error=err), 400
    fname = datetime.now(APP_TZ).strftime('database_%Y-%m-%d_%H-%M-%S.db')
    resp = Response(data, mimetype='application/x-sqlite3')
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@app.route('/api/admin/import-database', methods=['POST'])
def import_database_file():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    raw = request.get_data()
    if not raw or len(raw) < 512:
        return jsonify(success=False, error='Слишком маленький файл или пустое тело запроса'), 400

    abs_path = os.path.abspath(DB_PATH)
    ddir = os.path.dirname(abs_path) or '.'
    staging = os.path.join(ddir, f'.import_db_{os.getpid()}_{threading.get_ident()}.tmp')
    auto_backup = None
    try:
        with open(staging, 'wb') as f:
            f.write(raw)

        ok, verr = _validate_portal_sqlite_file(staging)
        if not ok:
            return jsonify(success=False, error=verr or 'Файл не прошёл проверку'), 400

        ts = datetime.now(APP_TZ).strftime('%Y-%m-%d_%H-%M-%S')
        if os.path.isfile(abs_path):
            auto_backup = os.path.join(ddir, f'database_before_import_{ts}.db')
            src_live = sqlite3.connect(abs_path, timeout=30)
            try:
                bak_dst = sqlite3.connect(auto_backup)
                try:
                    src_live.backup(bak_dst)
                finally:
                    bak_dst.close()
            finally:
                src_live.close()

        _wal_checkpoint_truncate_and_remove_sidecars(abs_path)
        os.replace(staging, abs_path)
        staging = None
        _remove_db_wal_shm(abs_path)

        return jsonify(
            success=True,
            backup_file=os.path.basename(auto_backup) if auto_backup else None,
        )
    except PermissionError as exc:
        return jsonify(
            success=False,
            error=f'Нет доступа к файлу базы (возможно, файл занят другим процессом): {exc}',
        ), 503
    except OSError as exc:
        return jsonify(success=False, error=str(exc)), 500
    finally:
        if staging and os.path.isfile(staging):
            try:
                os.unlink(staging)
            except OSError:
                pass


@app.route('/api/admin/sync-users-group', methods=['POST'])
def sync_users_group_api():
    """Заполнить группу «Пользователи» учётными записями из AD."""
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    conn = get_db_connection()
    try:
        ensure_default_users_group(conn)
        added = sync_all_ad_users_to_default_group(conn)
        group_id = get_users_group_id(conn)
        total_members = 0
        if group_id:
            total_members = conn.execute(
                'SELECT COUNT(*) AS cnt FROM group_members WHERE group_id = ?',
                (group_id,),
            ).fetchone()['cnt']
        conn.commit()
    finally:
        conn.close()
    return jsonify(success=True, added=added, total_members=total_members)


@app.route('/manage_ai_access', methods=['POST'])
def manage_ai_access():
    if session.get('username') not in MASTER_ADMINS:
        return jsonify(success=False), 403
    payload = request.json or {}
    action = (payload.get('action') or '').strip()
    entity_type = (payload.get('type') or 'user').strip().lower()
    raw_login = payload.get('username')
    if entity_type == 'group':
        entity_login = str(raw_login or '').strip().lower()
    else:
        entity_type = 'user'
        entity_login = normalize_ad_username(raw_login)
    if action not in ('add', 'delete'):
        return jsonify(success=False, error='Неизвестное действие'), 400
    if not entity_login:
        return jsonify(success=False, error='Укажите пользователя или группу'), 400
    conn = get_db_connection()
    try:
        if action == 'add':
            conn.execute(
                'INSERT OR IGNORE INTO ai_privileged_entities (entity_type, entity_login) VALUES (?, ?)',
                (entity_type, entity_login)
            )
        else:
            conn.execute(
                'DELETE FROM ai_privileged_entities WHERE entity_type = ? AND entity_login = ?',
                (entity_type, entity_login)
            )
        conn.commit()
        return jsonify(success=True)
    finally:
        conn.close()


if __name__ == '__main__':
    os.makedirs(KNOWLEDGE_BASE_INSTRUCTIONS_DIR, exist_ok=True)
    init_db()
    _tabel_load_cache()
    _tabel_rebuild_index_from_cache()
    _schedule_tabel_scan(force=False)
    port = int(os.environ.get('PORT', '5004'))
    debug = os.environ.get('FLASK_DEBUG', '1') == '1'
    app.run(host='0.0.0.0', port=port, debug=debug, threaded=True)